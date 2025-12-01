from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file, Response, abort, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import plotly.graph_objs as go
import plotly.utils
import json
from werkzeug.utils import secure_filename
from sqlalchemy import func, text, case, or_
from sqlalchemy.orm import joinedload
from typing import Union, Tuple
from fpdf import FPDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.utils import ImageReader
from PIL import Image
import matplotlib.pyplot as plt
import io
import tempfile
from sqlalchemy.orm import relationship
from functools import wraps
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import zipfile
from threading import Thread

app = Flask(__name__)
app.config['SECRET_KEY'] = 'votre_cle_secrete_ici'

# Configuration de la base de données pour Render
if os.environ.get('DATABASE_URL'):
    # Utiliser la variable d'environnement PostgreSQL de Render
    database_url = os.environ.get('DATABASE_URL')
    if database_url and database_url.startswith('postgres://'):
        # Render utilise postgres:// mais SQLAlchemy attend postgresql://
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    print(f"✅ Configuration PostgreSQL détectée: {database_url[:50] if database_url else 'None'}...")
    print(f"   Type de base: PostgreSQL")
    print(f"   URL complète: {database_url}")
    
elif os.environ.get('RENDER'):
    # Sur Render sans DATABASE_URL, utiliser un chemin persistant pour SQLite (fallback uniquement)
    db_path = '/opt/render/project/src/ste_releve.db'
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
    print(f"⚠️ ATTENTION: Configuration SQLite Render (fallback) - DATABASE_URL manquant!")
    print(f"   Base de données: {db_path}")
    print(f"   ⚠️ Les données peuvent être perdues lors du redéploiement!")
    print(f"   🔧 Vérifiez que la base PostgreSQL 'ste-app-db' existe sur Render!")
else:
    # En local, utiliser le chemin relatif
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///ste_releve.db'
    print("Configuration SQLite locale")

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'  # type: ignore

# Vérifier que la base PostgreSQL est accessible (après la définition de db)
if os.environ.get('DATABASE_URL') and ('postgresql' in app.config['SQLALCHEMY_DATABASE_URI'] or 'postgres' in app.config['SQLALCHEMY_DATABASE_URI']):
    try:
        with app.app_context():
            db.engine.connect()
            print(f"   ✅ Connexion PostgreSQL réussie!")
    except Exception as e:
        print(f"   ❌ ERREUR: Impossible de se connecter à PostgreSQL: {e}")
        print(f"   🔧 Vérifiez que la base 'ste-app-db' existe sur Render!")
        print(f"   🔧 Vérifiez que DATABASE_URL est correctement configuré!")
        
# Créer les dossiers nécessaires s'ils n'existent pas
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'signatures'), exist_ok=True)
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'conges'), exist_ok=True)
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'personnel'), exist_ok=True)
# Dossier pour le template PDF
TEMPLATE_PDF_FOLDER = os.path.join('static', 'templates')
os.makedirs(TEMPLATE_PDF_FOLDER, exist_ok=True)
# Essayer d'abord avec underscores, puis avec tirets
TEMPLATE_PDF_PATH = os.path.join(TEMPLATE_PDF_FOLDER, 'formulaire_absence_vierge.pdf')
if not os.path.exists(TEMPLATE_PDF_PATH):
    TEMPLATE_PDF_PATH = os.path.join(TEMPLATE_PDF_FOLDER, 'formulaire-absence-vierge.pdf')

# Initialiser la base de données au démarrage (migrations automatiques)
# Cette fonction sera définie plus tard, mais on l'appelle ici pour les migrations
def _init_db_on_startup():
    """Initialise la base de données et exécute les migrations au démarrage"""
    try:
        with app.app_context():
            # Vérifier et ajouter la colonne is_manager si nécessaire
            try:
                inspector = db.inspect(db.engine)
                columns = [col['name'] for col in inspector.get_columns('user')]
                
                if 'is_manager' not in columns:
                    print("🔄 Migration: Ajout de la colonne is_manager...")
                    with db.engine.connect() as conn:
                        if 'sqlite' in str(db.engine.url):
                            conn.execute(text('ALTER TABLE user ADD COLUMN is_manager INTEGER DEFAULT 0'))
                        else:
                            # PostgreSQL : user est un mot réservé, il faut utiliser des guillemets
                            conn.execute(text('ALTER TABLE "user" ADD COLUMN is_manager BOOLEAN DEFAULT FALSE'))
                        conn.commit()
                    print("✅ Colonne is_manager ajoutée")
            except Exception as e:
                # Si la table user n'existe pas encore, on continue
                if 'no such table' not in str(e).lower():
                    print(f"⚠️ Migration is_manager: {e}")
            
            # Créer les tables de gestion du personnel si nécessaire
            try:
                inspector = db.inspect(db.engine)
                table_names = inspector.get_table_names()
                tables_to_create = ['personnel', 'working_days', 'leave_request', 'personnel_document', 'absence', 'formation', 'formation_document', 'manager_signature', 'leave_request_document']
                tables_missing = [t for t in tables_to_create if t not in table_names]
                
                if tables_missing:
                    print(f"🔄 Migration: Création des tables de gestion du personnel...")
                    db.create_all()
                    print(f"✅ Tables créées: {', '.join(tables_missing)}")
            except Exception as e:
                print(f"⚠️ Migration tables personnel: {e}")
            
            # Migration : ajouter la colonne site_id à la table personnel si nécessaire
            try:
                inspector = db.inspect(db.engine)
                table_names = inspector.get_table_names()
                if 'personnel' in table_names:
                    columns = [col['name'] for col in inspector.get_columns('personnel')]
                    if 'site_id' not in columns:
                        print("🔄 Migration: Ajout de la colonne site_id à la table personnel...")
                        with db.engine.connect() as conn:
                            if 'sqlite' in str(db.engine.url):
                                conn.execute(text('ALTER TABLE personnel ADD COLUMN site_id INTEGER'))
                            else:
                                conn.execute(text('ALTER TABLE personnel ADD COLUMN site_id INTEGER REFERENCES site(id)'))
                            conn.commit()
                        print("✅ Colonne site_id ajoutée à la table personnel")
            except Exception as e:
                if 'no such table' not in str(e).lower() and 'does not exist' not in str(e).lower():
                    print(f"⚠️ Migration site_id personnel: {e}")
            
            # Migration : ajouter la colonne societe à la table personnel si nécessaire
            try:
                inspector = db.inspect(db.engine)
                table_names = inspector.get_table_names()
                if 'personnel' in table_names:
                    columns = [col['name'] for col in inspector.get_columns('personnel')]
                    if 'societe' not in columns:
                        print("🔄 Migration: Ajout de la colonne societe à la table personnel...")
                        with db.engine.connect() as conn:
                            conn.execute(text('ALTER TABLE personnel ADD COLUMN societe VARCHAR(100)'))
                            conn.commit()
                        print("✅ Colonne societe ajoutée à la table personnel")
            except Exception as e:
                if 'no such table' not in str(e).lower() and 'does not exist' not in str(e).lower():
                    print(f"⚠️ Migration societe personnel: {e}")
            
            # Migration : ajouter la colonne statut à la table absence si nécessaire
            try:
                inspector = db.inspect(db.engine)
                if 'absence' in inspector.get_table_names():
                    columns = [col['name'] for col in inspector.get_columns('absence')]
                    
                    if 'statut' not in columns:
                        print("🔄 Migration: Ajout de la colonne statut à la table absence...")
                        try:
                            with db.engine.connect() as conn:
                                if 'sqlite' in str(db.engine.url):
                                    conn.execute(text('ALTER TABLE absence ADD COLUMN statut VARCHAR(20) DEFAULT "en_attente"'))
                                    conn.execute(text('UPDATE absence SET statut = "en_attente" WHERE statut IS NULL'))
                                else:
                                    conn.execute(text('ALTER TABLE absence ADD COLUMN statut VARCHAR(20) DEFAULT \'en_attente\''))
                                    conn.execute(text("UPDATE absence SET statut = 'en_attente' WHERE statut IS NULL"))
                                conn.commit()
                            print("✅ Colonne statut ajoutée à la table absence")
                        except Exception as e:
                            print(f"⚠️ Migration statut absence: {e}")
                    else:
                        # Mettre à jour les absences existantes qui n'ont pas de statut
                        try:
                            with db.engine.connect() as conn:
                                if 'sqlite' in str(db.engine.url):
                                    conn.execute(text('UPDATE absence SET statut = "en_attente" WHERE statut IS NULL'))
                                else:
                                    conn.execute(text("UPDATE absence SET statut = 'en_attente' WHERE statut IS NULL"))
                                conn.commit()
                        except Exception as e:
                            print(f"⚠️ Mise à jour statut absence existantes: {e}")
            except Exception as e:
                print(f"Erreur lors de la vérification de la colonne statut absence: {e}")
            
            # Migration : ajouter les colonnes type_formation et date_fin_validite à la table formation si nécessaire
            try:
                inspector = db.inspect(db.engine)
                if 'formation' in inspector.get_table_names():
                    columns = [col['name'] for col in inspector.get_columns('formation')]
                    
                    if 'type_formation' not in columns:
                        print("🔄 Migration: Ajout de la colonne type_formation à la table formation...")
                        try:
                            with db.engine.connect() as conn:
                                if 'sqlite' in str(db.engine.url):
                                    conn.execute(text('ALTER TABLE formation ADD COLUMN type_formation VARCHAR(20) DEFAULT "demande"'))
                                    conn.execute(text('UPDATE formation SET type_formation = "demande" WHERE type_formation IS NULL'))
                                else:
                                    conn.execute(text('ALTER TABLE formation ADD COLUMN type_formation VARCHAR(20) DEFAULT \'demande\''))
                                    conn.execute(text("UPDATE formation SET type_formation = 'demande' WHERE type_formation IS NULL"))
                                conn.commit()
                            print("✅ Colonne type_formation ajoutée à la table formation")
                        except Exception as e:
                            print(f"⚠️ Migration type_formation: {e}")
                    
                    if 'date_fin_validite' not in columns:
                        print("🔄 Migration: Ajout de la colonne date_fin_validite à la table formation...")
                        try:
                            with db.engine.connect() as conn:
                                conn.execute(text('ALTER TABLE formation ADD COLUMN date_fin_validite DATE'))
                                conn.commit()
                            print("✅ Colonne date_fin_validite ajoutée à la table formation")
                        except Exception as e:
                            print(f"⚠️ Migration date_fin_validite: {e}")
            except Exception as e:
                print(f"Erreur lors de la vérification des colonnes formation: {e}")
    except Exception as e:
        print(f"⚠️ Erreur lors de l'initialisation de la base de données: {e}")

# Exécuter les migrations au démarrage
_init_db_on_startup()

# Modèles de base de données
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)  # Augmenté de 120 à 255
    role = db.Column(db.String(20), default='operateur')  # operateur, chef_equipe, admin
    is_manager = db.Column(db.Boolean, default=False)  # Indique si l'utilisateur est manager pour la gestion du personnel
    page_accesses = relationship('UserPageAccess', back_populates='user', cascade='all, delete-orphan')
    personnel = relationship('Personnel', foreign_keys='Personnel.user_id', back_populates='user', uselist=False)

class Site(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(50), nullable=False)  # SMP ou LPZ
    description = db.Column(db.String(200))

class TypeReleve(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'), nullable=False)
    type_mesure = db.Column(db.String(20), nullable=False)  # totalisateur, basique, hebdomadaire
    unite = db.Column(db.String(20), nullable=False)
    frequence = db.Column(db.String(20), default='quotidien')  # quotidien, hebdomadaire
    jour_specifique = db.Column(db.String(20))  # lundi pour eau potable et coagulant

class Releve(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    # Index sur la date et le type pour accélérer les filtres fréquents
    date = db.Column(db.Date, nullable=False, index=True)
    type_releve_id = db.Column(db.Integer, db.ForeignKey('type_releve.id'), nullable=False, index=True)
    valeur = db.Column(db.Float, nullable=False)
    utilisateur_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    commentaire = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class PhotoReleve(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, index=True)
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'), nullable=False, index=True)
    nom_debitmetre = db.Column(db.String(100), nullable=False)
    fichier_photo = db.Column(db.String(200), nullable=False)  # Nom du fichier pour compatibilité
    contenu_photo = db.Column(db.LargeBinary, nullable=True)  # Contenu binaire de la photo
    utilisateur_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    commentaire = db.Column(db.Text)
    session_id = db.Column(db.String(50), nullable=False)  # Identifiant unique de la session de relevé
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Modèles pour les routines d'exploitation
class FormulaireRoutine(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), unique=True, nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class QuestionRoutine(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    formulaire_id = db.Column(db.Integer, db.ForeignKey('formulaire_routine.id'), nullable=False, index=True)
    id_question = db.Column(db.String(50), nullable=False, index=True)
    lieu = db.Column(db.String(100), nullable=False, index=True)
    question = db.Column(db.Text, nullable=False)
    ordre = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class ReponseRoutine(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    formulaire_id = db.Column(db.Integer, db.ForeignKey('formulaire_routine.id'), nullable=False, index=True)
    question_id = db.Column(db.Integer, db.ForeignKey('question_routine.id'), nullable=False, index=True)
    reponse = db.Column(db.String(20), nullable=False)  # 'Fait', 'Non Fait', 'Non Applicable'
    commentaire = db.Column(db.Text)
    date_creation = db.Column(db.Date, default=lambda: datetime.now().date(), index=True)
    heure_creation = db.Column(db.Time, default=lambda: datetime.now().time())
    utilisateur_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# Routes principales
@app.route('/')
@login_required
def index():
    today = datetime.now().date()
    # Statut relevés (join avec TypeReleve)
    smp_fait = db.session.query(Releve).join(TypeReleve).filter(TypeReleve.site_id == 1, Releve.date == today).count() > 0
    lpz_fait = db.session.query(Releve).join(TypeReleve).filter(TypeReleve.site_id == 2, Releve.date == today).count() > 0
    releves_status = {
        'SMP': smp_fait,
        'LPZ': lpz_fait
    }
    # Calcul régularité relevés
    releves_regularite = {}
    for nom, site_id in [('SMP', 1), ('LPZ', 2)]:
        type_releves = TypeReleve.query.filter_by(site_id=site_id).all()
        # On considère qu'un relevé est "fait" si au moins un relevé existe ce jour-là pour ce site
        # Trouver la première date de relevé pour ce site (hors reset)
        reset = RESET_REGULARITE.get(('releve', f'Relevé {nom}'))
        first = db.session.query(Releve.date).join(TypeReleve).filter(TypeReleve.site_id == site_id)
        if reset:
            first = first.filter(Releve.date >= reset)
        first = first.order_by(Releve.date.asc()).first()
        if first:
            date_debut = first[0]
            jours = (today - date_debut).days + 1
            total = db.session.query(Releve.date).join(TypeReleve).filter(TypeReleve.site_id == site_id, Releve.date >= date_debut).distinct().count()
            regularite = int(100 * total / jours) if jours > 0 else 0
        else:
            regularite = 0
        releves_regularite[nom] = regularite
    # Routines fixes
    routines_list = [
        'STE PRINCIPALE SMP', 'STE CAB SMP', 'STEP SMP',
        'STE PRINCIPALE LPZ', 'STE CAB LPZ', 'STEP LPZ'
    ]
    routines_status = {}
    routines_regularite = {}
    for nom in routines_list:
        formulaire = FormulaireRoutine.query.filter_by(nom=nom).first()
        fait = False
        regularite = 0
        if formulaire:
            fait = db.session.query(ReponseRoutine).filter(ReponseRoutine.formulaire_id == formulaire.id, ReponseRoutine.date_creation == today).count() > 0
            reset = RESET_REGULARITE.get(('routine', nom))
            first = db.session.query(ReponseRoutine.date_creation).filter(ReponseRoutine.formulaire_id == formulaire.id)
            if reset:
                first = first.filter(ReponseRoutine.date_creation >= reset)
            first = first.order_by(ReponseRoutine.date_creation.asc()).first()
            if first:
                date_debut = first[0]
                jours = (today - date_debut).days + 1
                total = db.session.query(ReponseRoutine.date_creation).filter(ReponseRoutine.formulaire_id == formulaire.id, ReponseRoutine.date_creation >= date_debut).distinct().count()
                regularite = int(100 * total / jours) if jours > 0 else 0
        routines_status[nom] = fait
        routines_regularite[nom] = regularite
    return render_template('index.html', releves_status=releves_status, releves_regularite=releves_regularite, routines_list=routines_list, routines_status=routines_status, routines_regularite=routines_regularite, user_role=current_user.role)

# Fonction utilitaire pour trouver la première page autorisée
PAGE_REDIRECTS = [
    ('releve_site_smp', lambda: url_for('releve_site', site_id=1)),
    ('releve_site_lpz', lambda: url_for('releve_site', site_id=2)),
    ('historique', lambda: url_for('historique')),
    ('indicateurs', lambda: url_for('indicateurs')),
    ('routines', lambda: url_for('routines')),
    ('utilisateurs', lambda: url_for('utilisateurs')),
    ('perso', lambda: url_for('perso'))
]

def first_allowed_page(user):
    if user.role == 'admin':
        return url_for('index')
    accesses = {a.page_name: a.can_access for a in user.page_accesses}
    for page, url_func in PAGE_REDIRECTS:
        if accesses.get(page):
            return url_func()
    return url_for('logout')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            # Redirection selon les droits
            accesses = {a.page_name: a.can_access for a in user.page_accesses}
            if user.role == 'admin' or accesses.get('index'):
                return redirect(url_for('index'))
            else:
                return redirect(first_allowed_page(user))
        else:
            flash('Nom d\'utilisateur ou mot de passe incorrect')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/releves')
@login_required
def releves():
    return render_template('releves.html')

@app.route('/releve/<int:site_id>')
@login_required
def releve_site(site_id):
    if site_id == 1:
        return render_template('releve_smp.html')
    elif site_id == 2:
        return render_template('releve_lpz.html')
    else:
        return "Site inconnu", 404

@app.route('/api/releve', methods=['POST'])
@login_required
def ajouter_releve():
    data = request.get_json()
    print(f"DEBUG /api/releve - Données reçues: {data}")
    
    # Utiliser la date fournie ou aujourd'hui
    if 'date' in data and data['date']:
        date_releve = datetime.strptime(data['date'], '%Y-%m-%d').date()
    else:
        date_releve = datetime.now().date()
    
    if 'id' in data and data['id']:
        # Modification d'un relevé existant
        print(f"DEBUG /api/releve - Modification du relevé ID: {data['id']}")
        releve = Releve.query.get(data['id'])
        if releve:
            print(f"DEBUG /api/releve - Relevé trouvé: date={releve.date}, valeur={releve.valeur}, type_id={releve.type_releve_id}")
            
            # Utiliser le type_releve_id fourni ou garder l'ancien si undefined
            type_releve_id = data.get('type_releve_id')
            if type_releve_id == 'undefined' or not type_releve_id:
                type_releve_id = releve.type_releve_id
                print(f"DEBUG /api/releve - Type_releve_id undefined, utilisation de l'ancien: {type_releve_id}")
            
            print(f"DEBUG /api/releve - Nouvelles valeurs: date={date_releve}, valeur={data['valeur']}, type_id={type_releve_id}")
            
            releve.valeur = data['valeur']
            releve.commentaire = data.get('commentaire', '')
            releve.date = date_releve
            releve.type_releve_id = type_releve_id
            releve.utilisateur_id = current_user.id
            db.session.commit()
            print(f"DEBUG /api/releve - Relevé mis à jour avec succès")
            backup_database()  # Sauvegarde automatique après modification
            return jsonify({'success': True})
        else:
            print(f"DEBUG /api/releve - Relevé non trouvé pour ID: {data['id']}")
            return jsonify({'success': False, 'message': 'Relevé non trouvé'}), 404
    # Sinon, comportement existant (création ou update par date/type)
    releve_existant = Releve.query.filter_by(
        type_releve_id=data['type_releve_id'],
        date=date_releve
    ).first()
    if releve_existant:
        releve_existant.valeur = data['valeur']
        releve_existant.commentaire = data.get('commentaire', '')
        releve_existant.utilisateur_id = current_user.id
    else:
        nouveau_releve = Releve(
            date=date_releve,
            type_releve_id=data['type_releve_id'],
            valeur=data['valeur'],
            commentaire=data.get('commentaire', ''),
            utilisateur_id=current_user.id
        )
        db.session.add(nouveau_releve)
    db.session.commit()
    backup_database()  # Sauvegarde automatique après création/modification
    return jsonify({'success': True})

@app.route('/valeurs')
@login_required
def valeurs():
    sites = Site.query.all()
    return render_template('valeurs.html', sites=sites)

@app.route('/historique')
@login_required
def historique():
    sites = Site.query.all()
    return render_template('historique.html', sites=sites)

@app.route('/api/historique/<int:site_id>')
@login_required
def get_historique(site_id):
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')

    # Utiliser une jointure avec User pour éviter un N+1 (une requête par relevé)
    query = (
        db.session.query(
            Releve,
            TypeReleve,
            User.username.label('utilisateur_username')
        )
        .join(TypeReleve, Releve.type_releve_id == TypeReleve.id)
        .outerjoin(User, Releve.utilisateur_id == User.id)
        .filter(TypeReleve.site_id == site_id)
    )
    
    if date_debut:
        query = query.filter(Releve.date >= datetime.strptime(date_debut, '%Y-%m-%d').date())
    if date_fin:
        query = query.filter(Releve.date <= datetime.strptime(date_fin, '%Y-%m-%d').date())
    # Tri : date décroissante, puis id de TypeReleve croissant (ordre métier)
    releves = query.order_by(Releve.date.desc(), TypeReleve.id.asc()).all()

    result = []
    for releve, type_releve, utilisateur_username in releves:
        result.append({
            'id': releve.id,
            'date': releve.date.strftime('%Y-%m-%d'),
            'type_releve': type_releve.nom,
            'valeur': releve.valeur,
            'unite': type_releve.unite,
            'commentaire': releve.commentaire,
            'utilisateur': utilisateur_username or 'Inconnu'
        })
    
    return jsonify(result)

@app.route('/export_excel/<int:site_id>')
@login_required
def export_excel(site_id):
    site = db.session.get(Site, site_id)
    if not site:
        return "Site non trouvé", 404
    
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    
    # Récupérer tous les types de relevé (débitmètres) pour ce site, triés par id
    types_releve = TypeReleve.query.filter_by(site_id=site_id).order_by(TypeReleve.id).all()
    noms_debitmetres = [tr.nom for tr in types_releve]
    id_debitmetres = {tr.id: tr.nom for tr in types_releve}
    unites = {tr.nom: tr.unite for tr in types_releve}
    
    # Récupérer tous les relevés de la période
    query = db.session.query(Releve).filter(Releve.type_releve_id.in_(id_debitmetres.keys()))
    if date_debut:
        query = query.filter(Releve.date >= datetime.strptime(date_debut, '%Y-%m-%d').date())
    if date_fin:
        query = query.filter(Releve.date <= datetime.strptime(date_fin, '%Y-%m-%d').date())
    releves = query.all()
    
    # Construire le pivot : {date: {nom_debitmetre: valeur}}
    donnees = {}
    for r in releves:
        d = r.date.strftime('%Y-%m-%d')
        nom = id_debitmetres[r.type_releve_id]
        if d not in donnees:
            donnees[d] = {}
        donnees[d][nom] = r.valeur
    
    # Toutes les dates de la période (même sans relevé)
    if date_debut and date_fin:
        d1 = datetime.strptime(date_debut, '%Y-%m-%d').date()
        d2 = datetime.strptime(date_fin, '%Y-%m-%d').date()
        toutes_les_dates = [(d1 + timedelta(days=i)).strftime('%Y-%m-%d') for i in range((d2-d1).days+1)]
    else:
        toutes_les_dates = sorted(donnees.keys())
    
    # Créer le fichier Excel
    wb = Workbook()
    ws = wb.active
    if ws:
        ws.title = f"Relevés {site.nom}"
        
        # En-têtes : Date + noms de débitmètres
        ws.cell(row=1, column=1, value='Date').font = Font(bold=True)
        for i, nom in enumerate(noms_debitmetres, 2):
            cell = ws.cell(row=1, column=i, value=nom)
            cell.font = Font(bold=True)
        
        # Données
        for row, date in enumerate(toutes_les_dates, 2):
            ws.cell(row=row, column=1, value=date)
            for col, nom in enumerate(noms_debitmetres, 2):
                valeur = donnees.get(date, {}).get(nom, '')
                ws.cell(row=row, column=col, value=valeur)
    
    # Sauvegarder le fichier
    filename = f"releves_{site.nom}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb.save(filepath)
    
    print('DEBUG EXPORT EXCEL - Colonnes générées:', noms_debitmetres)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route('/indicateurs')
@login_required
def indicateurs():
    sites = Site.query.all()
    return render_template('indicateurs.html', sites=sites)

@app.route('/api/indicateurs/<int:site_id>')
@login_required
def get_indicateurs(site_id):
    site = Site.query.get_or_404(site_id)
    jours = request.args.get('jours', 30, type=int)
    date_debut = datetime.now().date() - timedelta(days=jours)
    
    # Récupérer tous les types de relevés du site
    types_releve = TypeReleve.query.filter_by(site_id=site_id).all()
    
    result = []
    for type_releve in types_releve:
        if type_releve.type_mesure == 'totalisateur':
            # Calculer les débits journaliers pour les totalisateurs
            releves = Releve.query.filter_by(type_releve_id=type_releve.id).filter(
                Releve.date >= date_debut
            ).order_by(Releve.date).all()
            
            if len(releves) > 1:
                valeurs = []
                for i in range(1, len(releves)):
                    difference = releves[i].valeur - releves[i-1].valeur
                    valeurs.append({
                        'date': releves[i].date.strftime('%Y-%m-%d'),
                        'valeur': difference
                    })
                
                result.append({
                    'nom': type_releve.nom,
                    'unite': type_releve.unite,
                    'valeurs': valeurs
                })
        elif type_releve.nom == 'Eau potable':
            # Traitement spécial pour l'eau potable (hebdomadaire avec calcul de différence)
            releves = Releve.query.filter_by(type_releve_id=type_releve.id).filter(
                Releve.date >= date_debut
            ).order_by(Releve.date).all()
            
            if len(releves) > 1:
                valeurs = []
                for i in range(1, len(releves)):
                    difference = releves[i].valeur - releves[i-1].valeur
                    # Calculer le numéro de semaine
                    semaine = releves[i].date.isocalendar()
                    semaine_label = f"S{semaine[1]}-{semaine[0]}"
                    valeurs.append({
                        'date': semaine_label,
                        'valeur': difference
                    })
                
                result.append({
                    'nom': type_releve.nom,
                    'unite': type_releve.unite,
                    'valeurs': valeurs
                })
        elif type_releve.nom == 'Coagulant':
            # Traitement spécial pour le coagulant (hebdomadaire sans calcul)
            releves = Releve.query.filter_by(type_releve_id=type_releve.id).filter(
                Releve.date >= date_debut
            ).order_by(Releve.date).all()
            
            if releves:
                valeurs = []
                for releve in releves:
                    # Calculer le numéro de semaine
                    semaine = releve.date.isocalendar()
                    semaine_label = f"S{semaine[1]}-{semaine[0]}"
                    valeurs.append({
                        'date': semaine_label,
                        'valeur': releve.valeur
                    })
                
                result.append({
                    'nom': type_releve.nom,
                    'unite': type_releve.unite,
                    'valeurs': valeurs
                })
        else:
            # Relevés basiques
            releves = Releve.query.filter_by(type_releve_id=type_releve.id).filter(
                Releve.date >= date_debut
            ).order_by(Releve.date).all()
            
            if releves:
                valeurs = [{
                    'date': r.date.strftime('%Y-%m-%d'),
                    'valeur': r.valeur
                } for r in releves]
                
                result.append({
                    'nom': type_releve.nom,
                    'unite': type_releve.unite,
                    'valeurs': valeurs
                })
    
    return jsonify(result)

@app.route('/releve_20')
@login_required
def releve_20():
    sites = Site.query.all()
    return render_template('releve_20.html', sites=sites)

@app.route('/api/upload_photo', methods=['POST'])
@login_required
def upload_photo():
    if 'photo' not in request.files:
        return jsonify({'success': False, 'message': 'Aucun fichier sélectionné'}), 400
    
    file = request.files['photo']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Aucun fichier sélectionné'}), 400
    
    site_nom = request.form.get('site_id')  # Renommé pour clarifier
    nom_debitmetre = request.form.get('nom_debitmetre')
    commentaire = request.form.get('commentaire', '')
    
    if not all([site_nom, nom_debitmetre]):
        return jsonify({'success': False, 'message': 'Paramètres manquants'}), 400
    
    try:
        # Convertir le nom du site en ID
        site = Site.query.filter_by(nom=site_nom).first()
        if not site:
            return jsonify({'success': False, 'message': f'Site {site_nom} non trouvé'}), 400
        
        site_id = site.id  # ID numérique du site
        
        # Générer un nom de fichier unique
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{secure_filename(file.filename or 'photo.jpg')}"
        
        # Utiliser le session_id envoyé par le frontend ou en générer un par défaut
        session_id = request.form.get('session_id')
        if not session_id:
            session_id = f"{current_user.id}_{site_nom}_{timestamp}"
        
        # Lire le contenu du fichier
        file_content = file.read()
        file.seek(0)  # Remettre le curseur au début pour la sauvegarde fichier
        
        # Sauvegarder le fichier (pour compatibilité locale)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Enregistrer en base de données
        photo = PhotoReleve(
            date=datetime.now().date(),
            site_id=site_id,  # Utiliser l'ID numérique
            nom_debitmetre=nom_debitmetre,
            fichier_photo=filename,
            contenu_photo=file_content,  # Stocker le contenu binaire
            utilisateur_id=current_user.id,
            commentaire=commentaire,
            session_id=session_id
        )
        
        db.session.add(photo)
        db.session.commit()
        
        # Log pour debug
        print(f"PHOTO ENREGISTREE: {{'site_id': {site_id}, 'site_nom': '{site_nom}', 'nom_debitmetre': '{nom_debitmetre}', 'utilisateur_id': {current_user.id}, 'date': {photo.date}, 'fichier_photo': '{filename}', 'session_id': '{session_id}'}}")
        
        return jsonify({'success': True, 'message': 'Photo enregistrée avec succès'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erreur lors de l\'enregistrement: {str(e)}'}), 500

@app.route('/api/releve_20_status')
@login_required
def get_releve_20_status():
    """Récupère le statut des relevés du 20 pour chaque site"""
    try:
        # Débitmètres à photographier par site
        debitmetres_smp = [
            'Exhaure 1', 'Exhaure 2', 'Exhaure 3', 'Exhaure 4', 
            'Retour dessableur', 'Retour Orage'
        ]
        debitmetres_lpz = [
            'Exhaure 1', 'Exhaure 2', 'Retour dessableur'
        ]
        
        # Récupérer les sites
        site_smp = Site.query.filter_by(nom='SMP').first()
        site_lpz = Site.query.filter_by(nom='LPZ').first()
        
        # Fonction pour vérifier le statut d'un débitmètre
        def get_debitmetre_status(site_id, nom_debitmetre):
            # Vérifier s'il y a une photo pour ce débitmètre aujourd'hui
            photo_aujourd_hui = PhotoReleve.query.filter_by(
                site_id=site_id,
                nom_debitmetre=nom_debitmetre,
                date=datetime.now().date()
            ).first()
            
            if photo_aujourd_hui:
                return 'Terminé'
            else:
                # Vérifier s'il y a des photos récentes (ce mois)
                debut_mois = datetime.now().replace(day=1).date()
                photo_ce_mois = PhotoReleve.query.filter_by(
                    site_id=site_id,
                    nom_debitmetre=nom_debitmetre
                ).filter(PhotoReleve.date >= debut_mois).first()
                
                if photo_ce_mois:
                    return 'En cours'
                else:
                    return 'En attente'
        
        # Statut SMP
        smp_status = []
        if site_smp:
            for debitmetre in debitmetres_smp:
                smp_status.append({
                    'nom': debitmetre,
                    'statut': get_debitmetre_status(site_smp.id, debitmetre)
                })
        
        # Statut LPZ
        lpz_status = []
        if site_lpz:
            for debitmetre in debitmetres_lpz:
                lpz_status.append({
                    'nom': debitmetre,
                    'statut': get_debitmetre_status(site_lpz.id, debitmetre)
                })
        
        return jsonify({
            'success': True,
            'smp': smp_status,
            'lpz': lpz_status
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/photos')
@login_required
def photos():
    photos = PhotoReleve.query.order_by(PhotoReleve.date.desc()).all()
    return render_template('photos.html', photos=photos)

@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename))

@app.route('/api/liste_releves_20')
@login_required
def liste_releves_20():
    print("=== DIAGNOSTIC liste_releves_20 ===")
    
    # Récupérer toutes les photos, groupées par session_id (pas seulement aujourd'hui)
    photos = PhotoReleve.query.order_by(PhotoReleve.date.desc()).all()
    
    # Grouper par session_id
    sessions = {}
    for photo in photos:
        if photo.session_id not in sessions:
            sessions[photo.session_id] = {
                'date': photo.date,
                'site_id': photo.site_id,
                'utilisateur_id': photo.utilisateur_id,
                'photos': []
            }
        sessions[photo.session_id]['photos'].append(photo)
    
    print(f"Résultat du groupement par session: {[(session_id, data['date'], data['site_id'], data['utilisateur_id'], len(data['photos'])) for session_id, data in sessions.items()]}")
    
    result = []
    for session_id, data in sessions.items():
        print(f"Traitement session: {session_id}, date={data['date']}, site_id={data['site_id']}, utilisateur_id={data['utilisateur_id']}, nb_photos={len(data['photos'])}")
        
        # Récupérer les informations du site et de l'utilisateur
        # Gérer les deux formats : site_id comme entier ou comme chaîne
        site = None
        if isinstance(data['site_id'], int):
            # Nouveau format : site_id est un entier
            site = db.session.get(Site, data['site_id'])
        else:
            # Ancien format : site_id est une chaîne (nom du site)
            site = Site.query.filter_by(nom=data['site_id']).first()
        
        user = db.session.get(User, data['utilisateur_id'])
        
        if site and user:
            result.append({
                'session_id': session_id,
                'date': data['date'].strftime('%d/%m/%Y'),
                'date_iso': data['date'].strftime('%Y-%m-%d'),
                'site': site.nom,
                'site_id': data['site_id'],
                'utilisateur': user.username,
                'utilisateur_id': data['utilisateur_id'],
                'nb_photos': len(data['photos'])
            })
    
    print(f"Relevés finaux: {result}")
    print("=== FIN DIAGNOSTIC ===")
    return jsonify(result)

@app.route('/api/photos_releve_20')
@login_required
def photos_releve_20():
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({'error': 'Paramètre session_id manquant'}), 400
    
    photos = PhotoReleve.query.filter_by(session_id=session_id).all()
    result = []
    for p in photos:
        user = User.query.get(p.utilisateur_id)
        result.append({
            'nom_debitmetre': p.nom_debitmetre,
            'fichier_photo': p.fichier_photo,
            'commentaire': p.commentaire,
            'date': p.date.strftime('%d/%m/%Y'),
            'utilisateur': user.username if user else p.utilisateur_id
        })
    return jsonify(result)

@app.route('/api/supprimer_releve_20', methods=['DELETE'])
@login_required
def supprimer_releve_20():
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({'success': False, 'message': 'Paramètre session_id manquant'}), 400
    
    try:
        # Récupérer toutes les photos de cette session
        photos = PhotoReleve.query.filter_by(session_id=session_id).all()
        
        # Supprimer les fichiers physiques
        for photo in photos:
            try:
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], photo.fichier_photo)
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                print(f"Erreur lors de la suppression du fichier {photo.fichier_photo}: {e}")
        
        # Supprimer les enregistrements de la base
        PhotoReleve.query.filter_by(session_id=session_id).delete()
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Relevé supprimé avec succès ({len(photos)} photos)'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erreur lors de la suppression: {str(e)}'}), 500

@app.route('/api/veille_releve_20/<site_id>')
@login_required
def veille_releve_20(site_id):
    # Récupérer la date d'hier
    hier = datetime.now().date() - timedelta(days=1)
    # Récupérer tous les types de relevé pour ce site
    types = TypeReleve.query.filter_by(site_id=Site.query.filter_by(nom=site_id).first().id).all()
    result = {}
    for tr in types:
        # On ne prend que les débitmètres (totalisateur)
        if tr.type_mesure == 'totalisateur':
            releve_hier = Releve.query.filter_by(type_releve_id=tr.id, date=hier).first()
            result[tr.nom] = releve_hier.valeur if releve_hier else None
    return jsonify(result)

@app.route('/api/releves_jour/<int:site_id>')
@login_required
def api_releves_jour(site_id):
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'error': 'Date manquante'}), 400
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'error': 'Format de date invalide'}), 400
    types_releve = TypeReleve.query.filter_by(site_id=site_id).all()
    result = {}
    for tr in types_releve:
        releve = Releve.query.filter_by(type_releve_id=tr.id, date=date_obj).first()
        if releve:
            result[tr.id] = {
                'valeur': releve.valeur,
                'unite': tr.unite
            }
    return jsonify(result)

# Nouvelles routes API pour les relevés SMP et LPZ
@app.route('/api/releves_smp', methods=['GET', 'POST'])
@login_required
def api_releves_smp() -> Union[Response, Tuple[Response, int]]:
    if request.method == 'GET':
        date = request.args.get('date')
        if not date:
            return jsonify({'error': 'Date requise'}), 400
        
        try:
            date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Format de date invalide'}), 400
        
        # Récupérer tous les types de relevé pour SMP
        types_releve = TypeReleve.query.filter_by(site_id=1).order_by(TypeReleve.id).all()
        
        # Récupérer les relevés existants pour cette date
        releves_existants = Releve.query.join(TypeReleve).filter(
            TypeReleve.site_id == 1,
            Releve.date == date_obj
        ).all()
        
        # Créer un dictionnaire des relevés existants
        releves_dict = {r.type_releve_id: r for r in releves_existants}
        
        result = []
        for tr in types_releve:
            releve = releves_dict.get(tr.id)
            result.append({
                'id': tr.id,
                'nom': tr.nom,
                'type_mesure': tr.type_mesure,
                'unite': tr.unite,
                'frequence': tr.frequence,
                'jour_specifique': tr.jour_specifique,
                'valeur': releve.valeur if releve else None,
                'commentaire': releve.commentaire if releve else '',
                'releve_id': releve.id if releve else None
            })
        
        return jsonify(result)
    
    elif request.method == 'POST':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Données JSON requises'}), 400
        try:
            date_obj = datetime.strptime(data['date'], '%Y-%m-%d').date()
        except (KeyError, ValueError):
            return jsonify({'error': 'Date invalide'}), 400
        # Traiter chaque relevé
        for releve_data in data.get('releves', []):
            if not isinstance(releve_data, dict):
                continue  # sécurité : ignorer si ce n'est pas un dict
            type_releve_id = releve_data.get('type_releve_id')
            valeur = releve_data.get('valeur')
            if type_releve_id is None or valeur is None:
                continue
            releve_existant = Releve.query.filter_by(
                type_releve_id=type_releve_id,
                date=date_obj
            ).first()
            if releve_existant:
                releve_existant.valeur = valeur
                releve_existant.commentaire = releve_data.get('commentaire', '')
                releve_existant.utilisateur_id = current_user.id
            else:
                nouveau_releve = Releve(
                    date=date_obj,
                    type_releve_id=type_releve_id,
                    valeur=valeur,
                    commentaire=releve_data.get('commentaire', ''),
                    utilisateur_id=current_user.id
                )
                db.session.add(nouveau_releve)
        try:
            db.session.commit()
            backup_database()  # Sauvegarde automatique après sauvegarde SMP
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Erreur lors de la sauvegarde: {str(e)}'}), 500
    return jsonify({'error': 'Méthode non supportée'}), 405

@app.route('/api/releves_lpz', methods=['GET', 'POST'])
@login_required
def api_releves_lpz() -> Union[Response, Tuple[Response, int]]:
    if request.method == 'GET':
        date = request.args.get('date')
        if not date:
            return jsonify({'error': 'Date requise'}), 400
        
        try:
            date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Format de date invalide'}), 400
        
        # Récupérer tous les types de relevé pour LPZ
        types_releve = TypeReleve.query.filter_by(site_id=2).order_by(TypeReleve.id).all()
        
        # Récupérer les relevés existants pour cette date
        releves_existants = Releve.query.join(TypeReleve).filter(
            TypeReleve.site_id == 2,
            Releve.date == date_obj
        ).all()
        
        # Créer un dictionnaire des relevés existants
        releves_dict = {r.type_releve_id: r for r in releves_existants}
        
        result = []
        for tr in types_releve:
            releve = releves_dict.get(tr.id)
            result.append({
                'id': tr.id,
                'nom': tr.nom,
                'type_mesure': tr.type_mesure,
                'unite': tr.unite,
                'frequence': tr.frequence,
                'jour_specifique': tr.jour_specifique,
                'valeur': releve.valeur if releve else None,
                'commentaire': releve.commentaire if releve else '',
                'releve_id': releve.id if releve else None
            })
        
        return jsonify(result)
    
    elif request.method == 'POST':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Données JSON requises'}), 400
        try:
            date_obj = datetime.strptime(data['date'], '%Y-%m-%d').date()
        except (KeyError, ValueError):
            return jsonify({'error': 'Date invalide'}), 400
        for releve_data in data.get('releves', []):
            if not isinstance(releve_data, dict):
                continue
            type_releve_id = releve_data.get('type_releve_id')
            valeur = releve_data.get('valeur')
            if type_releve_id is None or valeur is None:
                continue
            releve_existant = Releve.query.filter_by(
                type_releve_id=type_releve_id,
                date=date_obj
            ).first()
            if releve_existant:
                releve_existant.valeur = valeur
                releve_existant.commentaire = releve_data.get('commentaire', '')
                releve_existant.utilisateur_id = current_user.id
            else:
                nouveau_releve = Releve(
                    date=date_obj,
                    type_releve_id=type_releve_id,
                    valeur=valeur,
                    commentaire=releve_data.get('commentaire', ''),
                    utilisateur_id=current_user.id
                )
                db.session.add(nouveau_releve)
        try:
            db.session.commit()
            backup_database()  # Sauvegarde automatique après sauvegarde LPZ
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Erreur lors de la sauvegarde: {str(e)}'}), 500
    return jsonify({'error': 'Méthode non supportée'}), 405

@app.route('/api/veille/<int:site_id>')
@login_required
def api_veille(site_id):
    # Récupérer les valeurs de la veille pour un site donné
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'error': 'Date requise'}), 400
    
    date_releve = datetime.strptime(date_str, '%Y-%m-%d').date()
    date_veille = date_releve - timedelta(days=1)
    
    # Récupérer tous les types de relevé pour ce site
    types_releve = TypeReleve.query.filter_by(site_id=site_id).all()
    
    # Récupérer les relevés de la veille
    releves_veille = {}
    for tr in types_releve:
        releve = Releve.query.filter_by(
            type_releve_id=tr.id,
            date=date_veille
        ).first()
        if releve:
            releves_veille[tr.nom] = releve.valeur
    
    return jsonify({
        'date_veille': date_veille.strftime('%Y-%m-%d'),
        'releves': releves_veille
    })

@app.route('/api/verifier_existence/<int:site_id>')
@login_required
def api_verifier_existence(site_id):
    # Vérifier si des relevés existent déjà pour une date donnée
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'error': 'Date requise'}), 400
    
    date_releve = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    # Récupérer tous les types de relevé pour ce site
    types_releve = TypeReleve.query.filter_by(site_id=site_id).all()
    
    # Vérifier les relevés existants
    releves_existants = []
    for tr in types_releve:
        releve = Releve.query.filter_by(
            type_releve_id=tr.id,
            date=date_releve
        ).first()
        if releve:
            releves_existants.append({
                'nom': tr.nom,
                'valeur': releve.valeur,
                'utilisateur': User.query.get(releve.utilisateur_id).username
            })
    
    return jsonify({
        'date': date_str,
        'existe': len(releves_existants) > 0,
        'releves_existants': releves_existants
    })

@app.route('/api/releve/<int:releve_id>', methods=['DELETE'])
@login_required
def supprimer_releve(releve_id):
    try:
        releve = Releve.query.get_or_404(releve_id)
        db.session.delete(releve)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Relevé supprimé avec succès'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erreur lors de la suppression: {str(e)}'}), 500

@app.route('/api/releves_jour/<int:site_id>', methods=['DELETE'])
@login_required
def supprimer_releves_jour(site_id):
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'success': False, 'message': 'Date requise'}), 400
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        types_releve = TypeReleve.query.filter_by(site_id=site_id).all()
        type_ids = [tr.id for tr in types_releve]
        Releve.query.filter(Releve.type_releve_id.in_(type_ids), Releve.date == date_obj).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Tous les relevés de la journée ont été supprimés'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erreur lors de la suppression groupée: {str(e)}'}), 500

@app.route('/api/statistiques/<int:site_id>')
@login_required
def get_statistiques(site_id):
    jours = request.args.get('jours', 30, type=int)
    date_debut = datetime.now().date() - timedelta(days=jours)
    types_releve = TypeReleve.query.filter_by(site_id=site_id).all()
    stats = []
    for tr in types_releve:
        releves = Releve.query.filter_by(type_releve_id=tr.id).filter(Releve.date >= date_debut).all()
        if not releves:
            continue
        valeurs = [r.valeur for r in releves]
        stat = {
            'nom': tr.nom,
            'moyenne': sum(valeurs)/len(valeurs) if valeurs else 0,
            'min': min(valeurs) if valeurs else 0,
            'max': max(valeurs) if valeurs else 0,
            'total': sum(valeurs) if valeurs else 0,
            'unite': tr.unite
        }
        stats.append(stat)
    return jsonify(stats)

@app.route('/api/types_releve/<int:site_id>')
@login_required
def api_types_releve(site_id):
    types = TypeReleve.query.filter_by(site_id=site_id).all()
    result = [
        {
            'id': tr.id,
            'nom': tr.nom,
            'unite': tr.unite,
            'type_mesure': tr.type_mesure,
            'frequence': tr.frequence,
            'jour_specifique': tr.jour_specifique
        }
        for tr in types
    ]
    return jsonify(result)

@app.route('/api/indicateurs_donnee/<int:type_releve_id>')
@login_required
def api_indicateurs_donnee(type_releve_id):
    jours = request.args.get('jours', 30, type=int)
    date_debut = datetime.now().date() - timedelta(days=jours)
    type_releve = TypeReleve.query.get_or_404(type_releve_id)
    try:
        if type_releve.type_mesure == 'totalisateur':
            releves = Releve.query.filter_by(type_releve_id=type_releve_id).filter(Releve.date >= date_debut).order_by(Releve.date).all()
            valeurs = []
            if len(releves) > 1:
                for i in range(1, len(releves)):
                    difference = releves[i].valeur - releves[i-1].valeur
                    valeurs.append({
                        'date': releves[i].date.strftime('%Y-%m-%d'),
                        'valeur': difference
                    })
            print(f"[API indicateurs_donnee] type={type_releve.nom} valeurs={valeurs}")
            return jsonify([{
                'nom': type_releve.nom,
                'unite': type_releve.unite,
                'valeurs': valeurs
            }])
        elif type_releve.nom == 'Eau potable':
            # Traitement spécial pour l'eau potable (hebdomadaire avec calcul de différence)
            releves = Releve.query.filter_by(type_releve_id=type_releve_id).filter(Releve.date >= date_debut).order_by(Releve.date).all()
            valeurs = []
            if len(releves) > 1:
                for i in range(1, len(releves)):
                    difference = releves[i].valeur - releves[i-1].valeur
                    # Calculer le numéro de semaine
                    semaine = releves[i].date.isocalendar()
                    semaine_label = f"S{semaine[1]}-{semaine[0]}"
                    valeurs.append({
                        'date': semaine_label,
                        'valeur': difference
                    })
            print(f"[API indicateurs_donnee] Eau potable - valeurs={valeurs}")
            return jsonify([{
                'nom': type_releve.nom,
                'unite': type_releve.unite,
                'valeurs': valeurs
            }])
        elif type_releve.nom == 'Coagulant':
            # Traitement spécial pour le coagulant (hebdomadaire sans calcul)
            releves = Releve.query.filter_by(type_releve_id=type_releve_id).filter(Releve.date >= date_debut).order_by(Releve.date).all()
            valeurs = []
            for releve in releves:
                # Calculer le numéro de semaine
                semaine = releve.date.isocalendar()
                semaine_label = f"S{semaine[1]}-{semaine[0]}"
                valeurs.append({
                    'date': semaine_label,
                    'valeur': releve.valeur
                })
            print(f"[API indicateurs_donnee] Coagulant - valeurs={valeurs}")
            return jsonify([{
                'nom': type_releve.nom,
                'unite': type_releve.unite,
                'valeurs': valeurs
            }])
        else:
            releves = Releve.query.filter_by(type_releve_id=type_releve_id).filter(Releve.date >= date_debut).order_by(Releve.date).all()
            valeurs = [
                {'date': r.date.strftime('%Y-%m-%d'), 'valeur': r.valeur}
                for r in releves
            ]
            print(f"[API indicateurs_donnee] type={type_releve.nom} valeurs={valeurs}")
            return jsonify([{
                'nom': type_releve.nom,
                'unite': type_releve.unite,
                'valeurs': valeurs
            }])
    except Exception as e:
        print(f"[API indicateurs_donnee] ERREUR: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/rapport_pdf')
@login_required
def rapport_pdf():
    # Récupérer les paramètres
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    sites_param = request.args.get('sites')  # ex: "SMP,LPZ"
    if not date_debut or not date_fin or not sites_param:
        return "Paramètres manquants", 400
    try:
        date_debut_dt = datetime.strptime(date_debut, '%Y-%m-%d').date()
        date_fin_dt = datetime.strptime(date_fin, '%Y-%m-%d').date()
    except Exception:
        return "Format de date invalide", 400
    site_noms = sites_param.split(',')
    sites = Site.query.filter(Site.nom.in_(site_noms)).all()
    if not sites:
        return "Aucun site trouvé", 400

    # Récupérer tous les types de relevé pour les sites sélectionnés
    types_releves = TypeReleve.query.filter(TypeReleve.site_id.in_([s.id for s in sites])).all()

    # Récupérer les données pour chaque type de relevé
    data_series = []
    for tr in types_releves:
        releves = Releve.query.filter_by(type_releve_id=tr.id).filter(
            Releve.date >= date_debut_dt, Releve.date <= date_fin_dt
        ).order_by(Releve.date).all()
        if not releves:
            continue
        
        if tr.type_mesure == 'totalisateur' and len(releves) > 1:
            valeurs = []
            for i in range(1, len(releves)):
                difference = releves[i].valeur - releves[i-1].valeur
                valeurs.append((releves[i].date, difference))
        elif tr.nom == 'Eau potable' and len(releves) > 1:
            # Traitement spécial pour l'eau potable
            valeurs = []
            for i in range(1, len(releves)):
                difference = releves[i].valeur - releves[i-1].valeur
                valeurs.append((releves[i].date, difference))
        elif tr.nom == 'Coagulant':
            # Traitement spécial pour le coagulant
            valeurs = [(r.date, r.valeur) for r in releves]
        else:
            valeurs = [(r.date, r.valeur) for r in releves]
        
        data_series.append({
            'nom': tr.nom,
            'site': Site.query.get(tr.site_id).nom,
            'unite': tr.unite,
            'valeurs': valeurs
        })

    # Générer les graphiques avec matplotlib et les stocker en mémoire
    images = []
    for serie in data_series:
        # Toujours générer un graphique, même si pas de valeurs
        if serie['valeurs']:
            dates = [d.strftime('%d/%m/%Y') for d, v in serie['valeurs']]
            valeurs = [v for d, v in serie['valeurs']]
        else:
            dates = []
            valeurs = []
        plt.figure(figsize=(6,3))
        if valeurs:
            plt.plot(dates, valeurs, marker='o')
        else:
            plt.text(0.5, 0.5, 'Aucune donnée pour cette période', ha='center', va='center', fontsize=12, color='red', transform=plt.gca().transAxes)
        plt.title(f"{serie['nom']} - {serie['site']}")
        plt.xlabel('Date')
        plt.ylabel(f"Valeur ({serie['unite']})")
        plt.xticks(rotation=45, fontsize=7)
        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        plt.close()
        buf.seek(0)
        images.append(buf.read())

    # Générer le PDF avec fpdf
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font('Arial', 'B', 16)
    pdf.add_page()
    pdf.cell(0, 10, f"Rapport des relevés STE", ln=1, align='C')
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f"Période : {date_debut} au {date_fin}", ln=1, align='C')
    pdf.cell(0, 10, f"Site(s) : {', '.join(site_noms)}", ln=1, align='C')
    pdf.ln(5)
    # 3 graphiques par page
    for i, img in enumerate(images):
        if i % 3 == 0 and i != 0:
            pdf.add_page()
        # Sauvegarder l'image temporairement
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_img:
            tmp_img.write(img)
            tmp_img_path = tmp_img.name
        pdf.image(tmp_img_path, x=15, y=pdf.get_y(), w=180)
        pdf.ln(65)
        os.remove(tmp_img_path)
    # Retourner le PDF
    pdf_out = io.BytesIO()
    pdf_bytes = pdf.output(dest='S')
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode('latin1')
    pdf_out.write(pdf_bytes)
    pdf_out.seek(0)
    return send_file(pdf_out, mimetype='application/pdf', as_attachment=False, download_name='rapport_releves.pdf')

@app.route('/attente_rapport_pdf')
def attente_rapport_pdf():
    return render_template('attente_rapport_pdf.html')

# Routes pour les routines d'exploitation
@app.route('/routines')
@login_required
def routines():
    return render_template('routines.html')

@app.route('/admin_routines')
@login_required
def admin_routines():
    if current_user.role != 'admin':
        flash('Accès non autorisé')
        return redirect(url_for('index'))
    return render_template('admin_routines.html')

@app.route('/remplir_routine/<int:formulaire_id>')
@login_required
def remplir_routine(formulaire_id):
    return render_template('remplir_routine.html', formulaire_id=formulaire_id)

@app.route('/recap_routines')
@login_required
def recap_routines():
    return render_template('recap_routines.html')

@app.route('/detail_routine/<int:formulaire_id>')
@login_required
def detail_routine(formulaire_id):
    return render_template('detail_routine.html', formulaire_id=formulaire_id)

# API Routes pour les routines
@app.route('/api/routines/formulaires')
@login_required
def api_formulaires():
    formulaires = FormulaireRoutine.query.order_by(FormulaireRoutine.nom).all()
    return jsonify([{
        'id': f.id,
        'nom': f.nom,
        'created_at': f.created_at.isoformat() if f.created_at else None
    } for f in formulaires])

@app.route('/api/routines/formulaires/<int:formulaire_id>/questions')
@login_required
def api_questions_formulaire(formulaire_id):
    questions = QuestionRoutine.query.filter_by(formulaire_id=formulaire_id).order_by(QuestionRoutine.lieu, QuestionRoutine.ordre, QuestionRoutine.id_question).all()
    return jsonify([{
        'id': q.id,
        'id_question': q.id_question,
        'lieu': q.lieu,
        'question': q.question,
        'ordre': q.ordre
    } for q in questions])

@app.route('/api/routines/import-excel', methods=['POST'])
@login_required
def api_import_excel():
    if current_user.role != 'admin':
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    
    file = request.files['file']
    formulaire_id = request.form.get('formulaireId')
    
    if not file or not formulaire_id:
        return jsonify({'error': 'Fichier et ID du formulaire requis'}), 400
    
    try:
        # Lire le fichier Excel
        df = pd.read_excel(file)
        
        if not all(col in df.columns for col in ['id', 'lieu', 'question']):
            return jsonify({'error': 'Le fichier doit contenir les colonnes: id, lieu, question'}), 400
        
        updated_count = 0
        inserted_count = 0
        
        for index, row in df.iterrows():
            if str(row['id']).strip() != '' and str(row['lieu']).strip() != '' and str(row['question']).strip() != '':
                id_question = str(row['id'])
                
                # Vérifier si la question existe déjà
                question_existante = QuestionRoutine.query.filter_by(
                    formulaire_id=formulaire_id,
                    id_question=id_question
                ).first()
                
                if question_existante:
                    # Mettre à jour
                    question_existante.lieu = str(row['lieu'])
                    question_existante.question = str(row['question'])
                    question_existante.ordre = int(str(index)) + 1
                    updated_count += 1
                else:
                    # Créer nouvelle question
                    nouvelle_question = QuestionRoutine(
                        formulaire_id=formulaire_id,
                        id_question=id_question,
                        lieu=str(row['lieu']),
                        question=str(row['question']),
                        ordre=int(str(index)) + 1
                    )
                    db.session.add(nouvelle_question)
                    inserted_count += 1
        
        db.session.commit()
        return jsonify({
            'message': 'Import réussi',
            'updated': updated_count,
            'inserted': inserted_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erreur lors de l\'import: {str(e)}'}), 500

@app.route('/api/routines/reponses', methods=['POST'])
@login_required
def api_sauvegarder_reponse():
    data = request.form.to_dict()
    formulaire_id = data.get('formulaireId')
    question_id = data.get('questionId')
    reponse = data.get('reponse')
    commentaire = data.get('commentaire', '')
    
    if not all([formulaire_id, question_id, reponse]):
        return jsonify({'error': 'Données manquantes'}), 400
    
    # Vérifier si une réponse existe déjà pour cette question aujourd'hui
    date_aujourdhui = datetime.now().date()
    reponse_existante = ReponseRoutine.query.filter_by(
        formulaire_id=formulaire_id,
        question_id=question_id,
        date_creation=date_aujourdhui
    ).first()
    
    if reponse_existante:
        # Mettre à jour la réponse existante
        reponse_existante.reponse = reponse
        reponse_existante.commentaire = commentaire
        reponse_existante.utilisateur_id = current_user.id
        reponse_existante.heure_creation = datetime.now().time()
        db.session.commit()
        
        return jsonify({
            'id': reponse_existante.id,
            'message': 'Réponse mise à jour',
            'updated': True
        })
    else:
        # Créer une nouvelle réponse
        nouvelle_reponse = ReponseRoutine(
            formulaire_id=formulaire_id,
            question_id=question_id,
            reponse=reponse,
            commentaire=commentaire,
            utilisateur_id=current_user.id,
            date_creation=date_aujourdhui,
            heure_creation=datetime.now().time()
        )
        
        db.session.add(nouvelle_reponse)
        db.session.commit()
        
        return jsonify({
            'id': nouvelle_reponse.id,
            'message': 'Réponse enregistrée',
            'updated': False
        })

@app.route('/api/routines/reponses/<date>')
@login_required
def api_reponses_date(date):
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400
    
    reponses = db.session.query(ReponseRoutine, QuestionRoutine, FormulaireRoutine).join(
        QuestionRoutine, ReponseRoutine.question_id == QuestionRoutine.id
    ).join(
        FormulaireRoutine, ReponseRoutine.formulaire_id == FormulaireRoutine.id
    ).filter(
        ReponseRoutine.date_creation == date_obj
    ).order_by(ReponseRoutine.heure_creation.desc()).all()
    
    result = []
    for reponse, question, formulaire in reponses:
        result.append({
            'id': reponse.id,
            'formulaire_id': reponse.formulaire_id,
            'formulaire_nom': formulaire.nom,
            'question_id': reponse.question_id,
            'id_question': question.id_question,
            'lieu': question.lieu,
            'question': question.question,
            'reponse': reponse.reponse,
            'commentaire': reponse.commentaire,
            'date_creation': reponse.date_creation.isoformat(),
            'heure_creation': reponse.heure_creation.isoformat() if reponse.heure_creation else None
        })
    
    return jsonify(result)

@app.route('/api/routines/reponses/<int:reponse_id>', methods=['PUT'])
@login_required
def api_modifier_reponse(reponse_id):
    reponse = ReponseRoutine.query.get(reponse_id)
    if not reponse:
        return jsonify({'error': 'Réponse non trouvée'}), 404
    # Vérifier que la réponse date d'aujourd'hui (utiliser UTC)
    if reponse.date_creation != datetime.utcnow().date():
        return jsonify({'error': 'Modification non autorisée'}), 403
    data = request.form.to_dict()
    reponse.reponse = data.get('reponse', reponse.reponse)
    reponse.commentaire = data.get('commentaire', reponse.commentaire)
    reponse.utilisateur_id = current_user.id
    db.session.commit()
    return jsonify({'message': 'Réponse modifiée'})

@app.route('/api/routines/reponses/<int:reponse_id>', methods=['DELETE'])
@login_required
def api_supprimer_reponse(reponse_id):
    reponse = ReponseRoutine.query.get(reponse_id)
    if not reponse:
        return jsonify({'error': 'Réponse non trouvée'}), 404
    
    # Vérifier que la réponse date d'aujourd'hui (utiliser UTC)
    if reponse.date_creation != datetime.utcnow().date():
        return jsonify({'error': 'Suppression non autorisée'}), 403
    
    db.session.delete(reponse)
    db.session.commit()
    return jsonify({'message': 'Réponse supprimée'})

@app.route('/api/routines/export-pdf/<date>')
@login_required
def api_export_pdf_routines(date):
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400
    
    reponses = db.session.query(ReponseRoutine, QuestionRoutine, FormulaireRoutine).join(
        QuestionRoutine, ReponseRoutine.question_id == QuestionRoutine.id
    ).join(
        FormulaireRoutine, ReponseRoutine.formulaire_id == FormulaireRoutine.id
    ).filter(
        ReponseRoutine.date_creation == date_obj
    ).order_by(FormulaireRoutine.nom, QuestionRoutine.lieu, QuestionRoutine.id_question).all()
    
    # Créer le PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Rapport des Routines STE', ln=1, align='C')
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f'Date: {date}', ln=1, align='C')
    pdf.ln(10)
    
    # Grouper par formulaire
    grouped_by_form = {}
    for reponse, question, formulaire in reponses:
        if formulaire.nom not in grouped_by_form:
            grouped_by_form[formulaire.nom] = []
        grouped_by_form[formulaire.nom].append((reponse, question))
    
    for form_name, form_reponses in grouped_by_form.items():
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, form_name, ln=1)
        pdf.ln(5)
        
        # Grouper par lieu
        grouped_by_lieu = {}
        for reponse, question in form_reponses:
            if question.lieu not in grouped_by_lieu:
                grouped_by_lieu[question.lieu] = []
            grouped_by_lieu[question.lieu].append((reponse, question))
        
        for lieu, lieu_reponses in grouped_by_lieu.items():
            pdf.set_font('Arial', '', 10)
            pdf.multi_cell(0, 5, f'Question {question.id_question}: {question.question}')
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 5, f'Réponse: {reponse.reponse}', ln=1)
            if reponse.commentaire:
                pdf.set_font('Arial', '', 9)
                pdf.multi_cell(0, 4, f'Commentaire: {reponse.commentaire}')
            pdf.ln(3)
        
        pdf.ln(5)
    
    # Retourner le PDF
    pdf_out = io.BytesIO()
    pdf_bytes = pdf.output(dest='S')
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode('latin1')
    pdf_out.write(pdf_bytes)
    pdf_out.seek(0)
    
    return send_file(
        pdf_out,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'rapport-routines-{date}.pdf'
    )

@app.route('/api/routines/export-excel/<date>')
@login_required
def api_export_excel_routines(date):
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400
    
    reponses = db.session.query(ReponseRoutine, QuestionRoutine, FormulaireRoutine).join(
        QuestionRoutine, ReponseRoutine.question_id == QuestionRoutine.id
    ).join(
        FormulaireRoutine, ReponseRoutine.formulaire_id == FormulaireRoutine.id
    ).filter(
        ReponseRoutine.date_creation == date_obj
    ).order_by(FormulaireRoutine.nom, QuestionRoutine.lieu, QuestionRoutine.id_question).all()
    
    # Créer le DataFrame
    data = []
    for reponse, question, formulaire in reponses:
        data.append({
            'Formulaire': formulaire.nom,
            'Lieu': question.lieu,
            'ID Question': question.id_question,
            'Question': question.question,
            'Réponse': reponse.reponse,
            'Commentaire': reponse.commentaire or '',
            'Heure': reponse.heure_creation.isoformat() if reponse.heure_creation else ''
        })
    
    df = pd.DataFrame(data)
    
    # Créer le fichier Excel temporaire
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        with pd.ExcelWriter(tmp_file.name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Rapport', index=False)
    
    return send_file(
        tmp_file.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'rapport-routines-{date}.xlsx'
    )

@app.route('/api/routines/stats/<date>')
@login_required
def api_stats_routines(date):
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400
    
    # Récupérer toutes les questions avec leurs réponses pour la date
    stats = db.session.query(
        QuestionRoutine.id_question,
        QuestionRoutine.question,
        QuestionRoutine.lieu,
        FormulaireRoutine.nom.label('formulaire_nom'),
        func.count(case((ReponseRoutine.reponse == 'Fait', 1))).label('fait'),
        func.count(case((ReponseRoutine.reponse == 'Non Fait', 1))).label('non_fait'),
        func.count(case((ReponseRoutine.reponse == 'Non Applicable', 1))).label('non_applicable'),
        func.count(ReponseRoutine.id).label('total')
    ).join(
        FormulaireRoutine, QuestionRoutine.formulaire_id == FormulaireRoutine.id
    ).outerjoin(
        ReponseRoutine, 
        db.and_(
            QuestionRoutine.id == ReponseRoutine.question_id,
            ReponseRoutine.date_creation == date_obj
        )
    ).group_by(
        QuestionRoutine.id,
        QuestionRoutine.id_question,
        QuestionRoutine.question,
        QuestionRoutine.lieu,
        FormulaireRoutine.nom
    ).order_by(
        FormulaireRoutine.nom,
        QuestionRoutine.lieu,
        QuestionRoutine.id_question
    ).all()
    
    result = []
    for stat in stats:
        result.append({
            'id_question': stat.id_question,
            'question': stat.question,
            'lieu': stat.lieu,
            'formulaire_nom': stat.formulaire_nom,
            'fait': stat.fait,
            'non_fait': stat.non_fait,
            'non_applicable': stat.non_applicable,
            'total': stat.total
        })
    
    return jsonify(result)

@app.route('/api/routines/export-excel/formulaire/<int:formulaire_id>')
@login_required
def api_export_excel_formulaire(formulaire_id):
    reponses = db.session.query(ReponseRoutine, QuestionRoutine, FormulaireRoutine).join(
        QuestionRoutine, ReponseRoutine.question_id == QuestionRoutine.id
    ).join(
        FormulaireRoutine, ReponseRoutine.formulaire_id == FormulaireRoutine.id
    ).filter(
        ReponseRoutine.formulaire_id == formulaire_id
    ).order_by(ReponseRoutine.date_creation.desc(), QuestionRoutine.lieu, QuestionRoutine.id_question).all()

    # Créer le DataFrame
    data = []
    for reponse, question, formulaire in reponses:
        data.append({
            'Date': reponse.date_creation.isoformat() if reponse.date_creation else '',
            'Heure': reponse.heure_creation.isoformat() if reponse.heure_creation else '',
            'Lieu': question.lieu,
            'ID Question': question.id_question,
            'Question': question.question,
            'Réponse': reponse.reponse,
            'Commentaire': reponse.commentaire or '',
            'Utilisateur': reponse.utilisateur_id,
        })

    df = pd.DataFrame(data)

    # Créer le fichier Excel temporaire
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        with pd.ExcelWriter(tmp_file.name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Réponses', index=False)

    return send_file(
        tmp_file.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'export-formulaire-{formulaire_id}.xlsx'
    )

@app.route('/api/routines/formulaires_remplis_aujourdhui')
@login_required
def api_formulaires_remplis_aujourdhui():
    today = datetime.utcnow().date()  # Utiliser UTC pour éviter les problèmes de fuseau horaire
    count = db.session.query(ReponseRoutine.formulaire_id).filter(ReponseRoutine.date_creation == today).distinct().count()
    return jsonify({'formulaires_remplis': count})

@app.route('/api/routines/reponses/<int:formulaire_id>/<date>', methods=['DELETE'])
@login_required
def api_supprimer_routine_journee(formulaire_id, date):
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400
    
    reponses = ReponseRoutine.query.filter_by(formulaire_id=formulaire_id, date_creation=date_obj).all()
    if not reponses:
        return jsonify({'error': 'Aucune réponse à supprimer'}), 404
    
    for rep in reponses:
        db.session.delete(rep)
    db.session.commit()
    
    return jsonify({'success': True})

# Initialisation de la base de données
def init_db():
    with app.app_context():
        db.create_all()
        
        # Migration : ajouter session_id aux photos existantes si nécessaire
        try:
            # Vérifier si la colonne session_id existe
            inspector = db.inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('photo_releve')]
            
            if 'session_id' not in columns:
                print("Migration : ajout de la colonne session_id...")
                try:
                    with db.engine.connect() as conn:
                        conn.execute(text('ALTER TABLE photo_releve ADD COLUMN session_id VARCHAR(50)'))
                        conn.commit()
                    
                    # Générer des session_id pour les photos existantes
                    photos = PhotoReleve.query.all()
                    for photo in photos:
                        timestamp = photo.created_at.strftime('%Y%m%d_%H%M%S') if photo.created_at else datetime.now().strftime('%Y%m%d_%H%M%S')
                        session_id = f"{photo.utilisateur_id}_{photo.site_id}_{timestamp}"
                        photo.session_id = session_id
                    
                    db.session.commit()
                    print(f"Migration terminée : {len(photos)} photos mises à jour")
                except Exception as e:
                    print(f"Erreur lors de l'ajout de la colonne : {e}")
                    # Si la colonne existe déjà, on continue
                    pass
            else:
                print("Colonne session_id déjà présente")
        except Exception as e:
            print(f"Erreur lors de la vérification de la migration : {e}")
        
        # Migration : ajouter contenu_photo aux photos existantes si nécessaire
        try:
            inspector = db.inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('photo_releve')]
            
            if 'contenu_photo' not in columns:
                print("Migration : ajout de la colonne contenu_photo...")
                try:
                    with db.engine.connect() as conn:
                        conn.execute(text('ALTER TABLE photo_releve ADD COLUMN contenu_photo BYTEA'))
                        conn.commit()
                    print("Colonne contenu_photo ajoutée avec succès")
                except Exception as e:
                    print(f"Erreur lors de l'ajout de la colonne contenu_photo : {e}")
            else:
                print("Colonne contenu_photo déjà présente")
        except Exception as e:
            print(f"Erreur lors de la vérification de la migration contenu_photo : {e}")
        
        # Migration : créer la table code_magasin si elle n'existe pas
        try:
            inspector = db.inspect(db.engine)
            if 'code_magasin' not in inspector.get_table_names():
                print("Migration : création de la table code_magasin...")
                db.create_all()  # Cela va créer la table code_magasin
                print("Table code_magasin créée avec succès")
            else:
                print("Table code_magasin déjà présente")
        except Exception as e:
            print(f"Erreur lors de la vérification de la table code_magasin : {e}")
        
        # Migration : ajouter la colonne is_manager à la table user si nécessaire
        try:
            inspector = db.inspect(db.engine)
            columns = [col['name'] for col in inspector.get_columns('user')]
            
            if 'is_manager' not in columns:
                print("Migration : ajout de la colonne is_manager...")
                try:
                    with db.engine.connect() as conn:
                        # SQLite utilise INTEGER pour les booléens
                        if 'sqlite' in str(db.engine.url):
                            conn.execute(text('ALTER TABLE user ADD COLUMN is_manager INTEGER DEFAULT 0'))
                        else:
                            # PostgreSQL : user est un mot réservé, il faut utiliser des guillemets
                            conn.execute(text('ALTER TABLE "user" ADD COLUMN is_manager BOOLEAN DEFAULT FALSE'))
                        conn.commit()
                    print("Colonne is_manager ajoutée avec succès")
                except Exception as e:
                    print(f"Erreur lors de l'ajout de la colonne is_manager: {e}")
            else:
                print("Colonne is_manager déjà présente")
        except Exception as e:
            print(f"Erreur lors de la vérification de la colonne is_manager: {e}")
        
        # Migration : créer les tables pour la gestion du personnel si elles n'existent pas
        try:
            inspector = db.inspect(db.engine)
            table_names = inspector.get_table_names()
            
            tables_to_create = ['personnel', 'working_days', 'leave_request', 'personnel_document', 'absence', 'formation', 'formation_document', 'manager_signature', 'leave_request_document']
            tables_missing = [t for t in tables_to_create if t not in table_names]
            
            if tables_missing:
                print(f"Migration : création des tables de gestion du personnel: {', '.join(tables_missing)}...")
                db.create_all()  # Cela va créer toutes les tables manquantes
                print(f"Tables créées avec succès: {', '.join(tables_missing)}")
            else:
                print("Tables de gestion du personnel déjà présentes")
        except Exception as e:
            print(f"Erreur lors de la vérification des tables de gestion du personnel: {e}")
        
        # Créer les sites
        if not Site.query.first():
            smp = Site(nom='SMP', description='Station de traitement des eaux SMP')
            lpz = Site(nom='LPZ', description='Station de traitement des eaux LPZ')
            db.session.add(smp)
            db.session.add(lpz)
            db.session.commit()
            
            # Créer les types de relevés pour SMP
            types_smp = [
                ('Exhaure 1', 'totalisateur', 'm³', 'quotidien'),
                ('Exhaure 2', 'totalisateur', 'm³', 'quotidien'),
                ('Exhaure 3', 'totalisateur', 'm³', 'quotidien'),
                ('Exhaure 4', 'totalisateur', 'm³', 'quotidien'),
                ('Retour dessableur', 'totalisateur', 'm³', 'quotidien'),
                ('Retour Orage', 'totalisateur', 'm³', 'quotidien'),
                ('Rejet à l\'Arc', 'totalisateur', 'm³', 'quotidien'),
                ('Surpresseur 4 pompes', 'totalisateur', 'm³', 'quotidien'),
                ('Surpresseur 7 pompes', 'totalisateur', 'm³', 'quotidien'),
                ('Entrée STE CAB', 'totalisateur', 'm³', 'quotidien'),
                ('Alimentation CAB', 'totalisateur', 'm³', 'quotidien'),
                ('Eau potable', 'totalisateur', 'm³', 'hebdomadaire', 'lundi'),
                ('Forage', 'totalisateur', 'm³', 'quotidien'),
                ('Boue STE', 'basique', 'pressées', 'quotidien'),
                ('Boue STE CAB', 'basique', 'pressées', 'quotidien'),
                ('pH entrée', 'basique', '', 'quotidien'),
                ('pH sortie', 'basique', '', 'quotidien'),
                ('Température entrée', 'basique', '°C', 'quotidien'),
                ('Température sortie', 'basique', '°C', 'quotidien'),
                ('Conductivité sortie', 'basique', 'µS/cm', 'quotidien'),
                ('MES entrée', 'basique', 'mg/L', 'quotidien'),
                ('MES sortie', 'basique', 'mg/L', 'quotidien'),
                ('Coagulant', 'basique', 'L', 'hebdomadaire', 'lundi'),
                ('Floculant', 'basique', 'kg', 'quotidien'),
                ('CO2', 'basique', '%', 'quotidien')
            ]
            
            for nom, type_mesure, unite, frequence, *args in types_smp:
                jour_specifique = args[0] if args else None
                tr = TypeReleve(
                    nom=nom,
                    site_id=smp.id,
                    type_mesure=type_mesure,
                    unite=unite,
                    frequence=frequence,
                    jour_specifique=jour_specifique
                )
                db.session.add(tr)
            
            # Créer les types de relevés pour LPZ
            types_lpz = [
                ('Exhaure 1', 'totalisateur', 'm³', 'quotidien'),
                ('Exhaure 2', 'totalisateur', 'm³', 'quotidien'),
                ('Retour dessableur', 'totalisateur', 'm³', 'quotidien'),
                ('Surpresseur BP', 'totalisateur', 'm³', 'quotidien'),
                ('Surpresseur HP', 'totalisateur', 'm³', 'quotidien'),
                ('Rejet à l\'Arc', 'totalisateur', 'm³', 'quotidien'),
                ('Entrée STE CAB', 'totalisateur', 'm³', 'quotidien'),
                ('Alimentation CAB', 'totalisateur', 'm³', 'quotidien'),
                ('Eau de montagne', 'totalisateur', 'm³', 'quotidien'),
                ('Eau potable', 'totalisateur', 'm³', 'hebdomadaire', 'lundi'),
                ('Boue STE', 'basique', 'pressées', 'quotidien'),
                ('Boue STE CAB', 'basique', 'pressées', 'quotidien'),
                ('pH entrée', 'basique', '', 'quotidien'),
                ('pH sortie', 'basique', '', 'quotidien'),
                ('Température entrée', 'basique', '°C', 'quotidien'),
                ('Température sortie', 'basique', '°C', 'quotidien'),
                ('Conductivité sortie', 'basique', 'µS/cm', 'quotidien'),
                ('MES entrée', 'basique', 'mg/L', 'quotidien'),
                ('MES sortie', 'basique', 'mg/L', 'quotidien'),
                ('Coagulant', 'basique', 'L', 'hebdomadaire', 'lundi'),
                ('Floculant', 'basique', 'kg', 'quotidien'),
                ('CO2', 'basique', '%', 'quotidien')
            ]
            
            for nom, type_mesure, unite, frequence, *args in types_lpz:
                jour_specifique = args[0] if args else None
                tr = TypeReleve(
                    nom=nom,
                    site_id=lpz.id,
                    type_mesure=type_mesure,
                    unite=unite,
                    frequence=frequence,
                    jour_specifique=jour_specifique
                )
                db.session.add(tr)
            
            # Créer l'utilisateur admin s'il n'existe pas
            admin = User.query.filter_by(username='admin').first()
            if not admin:
                admin = User()
                admin.username = 'admin'
                admin.password_hash = generate_password_hash('admin123')
                admin.role = 'admin'
                db.session.add(admin)
                print("✅ Utilisateur admin créé avec succès!")
                print("   Username: admin")
                print("   Password: admin123")
                print("   Role: admin")
            
            db.session.commit()

        # Initialiser les formulaires de routines par défaut
        formulaires_routines = [
            'STE PRINCIPALE LPZ', 'STE CAB LPZ', 'STEP LPZ', 
            'STE PRINCIPALE SMP', 'STE CAB SMP', 'STEP SMP'
        ]
        
        for nom in formulaires_routines:
            formulaire_existant = FormulaireRoutine.query.filter_by(nom=nom).first()
            if not formulaire_existant:
                nouveau_formulaire = FormulaireRoutine(nom=nom)
                db.session.add(nouveau_formulaire)
        
        db.session.commit()

        # Créer la configuration email par défaut
        email_config = EmailConfig.query.first()
        if not email_config:
            email_config = EmailConfig(
                email_address='admin@ste-releve.com',
                smtp_server='smtp.gmail.com',
                smtp_port=587
            )
            db.session.add(email_config)
            db.session.commit()
        
        # Mise à jour des unités pour les types de relevés existants
        print("🔄 Mise à jour des unités des types de relevés...")
        
        # Mise à jour pour SMP
        smp_types_to_update = {
            'Boue STE': 'pressées',
            'Boue STE CAB': 'pressées',
            'Floculant': 'kg',
            'CO2': '%'
        }
        
        for nom, nouvelle_unite in smp_types_to_update.items():
            type_releve = TypeReleve.query.filter_by(nom=nom, site_id=1).first()
            if type_releve and type_releve.unite != nouvelle_unite:
                print(f"📝 Mise à jour {nom} SMP: {type_releve.unite} → {nouvelle_unite}")
                type_releve.unite = nouvelle_unite
        
        # Mise à jour pour LPZ
        lpz_types_to_update = {
            'Boue STE': 'pressées',
            'Boue STE CAB': 'pressées',
            'Floculant': 'kg',
            'CO2': '%'
        }
        
        for nom, nouvelle_unite in lpz_types_to_update.items():
            type_releve = TypeReleve.query.filter_by(nom=nom, site_id=2).first()
            if type_releve and type_releve.unite != nouvelle_unite:
                print(f"📝 Mise à jour {nom} LPZ: {type_releve.unite} → {nouvelle_unite}")
                type_releve.unite = nouvelle_unite
        
        db.session.commit()
        print("✅ Mise à jour des unités terminée")
        print("✅ Base de données initialisée avec succès!")

# Liste des pages gérables pour les droits
PAGE_NAMES = [
    'index',
    'releve_site_smp',
    'releve_site_lpz',
    'historique',
    'indicateurs',
    'routines',
    'code_magasin',
    'utilisateurs',  # admin only
    'perso'  # gestion du personnel
]

# Modèle pour les droits d'accès par page
class UserPageAccess(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    page_name = db.Column(db.String(50), nullable=False)
    can_access = db.Column(db.Boolean, default=False)
    user = relationship('User', back_populates='page_accesses')

User.page_accesses = relationship('UserPageAccess', back_populates='user', cascade='all, delete-orphan')

# Décorateur pour vérifier l'accès à une page
def require_page_access(page_name):
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated_function(*args, **kwargs):
            if current_user.role == 'admin':
                # L'admin a toujours accès à tout
                return f(*args, **kwargs)
            access = next((a for a in current_user.page_accesses if a.page_name == page_name), None)
            if access and access.can_access:
                return f(*args, **kwargs)
            abort(403)
        return decorated_function
    return decorator

def generate_daily_code():
    """Génère un code de 4 chiffres pour aujourd'hui"""
    import random
    aujourd_hui = datetime.now().date()
    
    # Vérifier si un code existe déjà pour aujourd'hui
    code_existant = CodeMagasin.query.filter_by(date=aujourd_hui).first()
    if code_existant:
        return code_existant.code
    
    # Générer un nouveau code de 4 chiffres
    code = str(random.randint(1000, 9999))
    
    # Créer et sauvegarder le nouveau code
    nouveau_code = CodeMagasin(code=code, date=aujourd_hui)
    db.session.add(nouveau_code)
    db.session.commit()
    
    return code

@app.route('/code_magasin')
@require_page_access('code_magasin')
def code_magasin():
    # Générer le code du jour
    code_actuel = generate_daily_code()
    
    # Récupérer les codes du mois en cours
    aujourd_hui = datetime.now().date()
    premier_jour_mois = aujourd_hui.replace(day=1)
    dernier_jour_mois = (premier_jour_mois.replace(month=premier_jour_mois.month % 12 + 1, day=1) - timedelta(days=1)) if premier_jour_mois.month < 12 else premier_jour_mois.replace(year=premier_jour_mois.year + 1, month=1, day=1) - timedelta(days=1)
    
    codes_mensuels = CodeMagasin.query.filter(
        CodeMagasin.date >= premier_jour_mois,
        CodeMagasin.date <= dernier_jour_mois
    ).order_by(CodeMagasin.date.desc()).all()
    
    # Dictionnaire pour traduire les mois en français
    mois_francais = {
        1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
        5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
        9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
    }
    
    mois_annee = f"{mois_francais[aujourd_hui.month]} {aujourd_hui.year}"
    
    return render_template('code_magasin.html', 
                         current_code=code_actuel, 
                         monthly_codes=codes_mensuels,
                         current_month=mois_annee,
                         today=aujourd_hui)

@app.route('/api/code_magasin/current')
@require_page_access('code_magasin')
def api_current_code():
    """API pour récupérer le code actuel"""
    code_actuel = generate_daily_code()
    return jsonify({
        'code': code_actuel,
        'date': datetime.now().date().isoformat()
    })

@app.route('/api/code_magasin/monthly')
@require_page_access('code_magasin')
def api_monthly_codes():
    """API pour récupérer les codes du mois"""
    aujourd_hui = datetime.now().date()
    premier_jour_mois = aujourd_hui.replace(day=1)
    dernier_jour_mois = (premier_jour_mois.replace(month=premier_jour_mois.month % 12 + 1, day=1) - timedelta(days=1)) if premier_jour_mois.month < 12 else premier_jour_mois.replace(year=premier_jour_mois.year + 1, month=1, day=1) - timedelta(days=1)
    
    codes_mensuels = CodeMagasin.query.filter(
        CodeMagasin.date >= premier_jour_mois,
        CodeMagasin.date <= dernier_jour_mois
    ).order_by(CodeMagasin.date.desc()).all()
    
    donnees_codes = []
    for code_entry in codes_mensuels:
        donnees_codes.append({
            'date': code_entry.date.strftime('%d/%m/%Y'),
            'code': code_entry.code,
            'is_today': code_entry.date == aujourd_hui
        })
    
    # Dictionnaire pour traduire les mois en français
    mois_francais = {
        1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
        5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
        9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
    }
    
    mois_annee = f"{mois_francais[aujourd_hui.month]} {aujourd_hui.year}"
    
    return jsonify({
        'codes': donnees_codes,
        'month': mois_annee
    })

@app.route('/utilisateurs')
@require_page_access('utilisateurs')
def utilisateurs():
    users = User.query.all()
    page_names = PAGE_NAMES
    return render_template('utilisateurs.html', users=users, page_names=page_names)

@app.route('/api/utilisateurs', methods=['GET', 'POST'])
@require_page_access('utilisateurs')
def api_utilisateurs():
    if request.method == 'GET':
        users = User.query.all()
        data = []
        for user in users:
            data.append({
                'id': user.id,
                'username': user.username,
                'role': user.role,
                'is_manager': user.is_manager,
                'page_accesses': {a.page_name: a.can_access for a in user.page_accesses}
            })
        return jsonify(data)
    elif request.method == 'POST':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Données JSON manquantes'}), 400
        username = data.get('username')
        password = data.get('password')
        role = data.get('role', 'operateur')
        is_manager = data.get('is_manager', False)
        if not username or not password:
            return jsonify({'error': 'Champs manquants'}), 400
        if User.query.filter_by(username=username).first():
            return jsonify({'error': 'Nom d\'utilisateur déjà pris'}), 400
        user = User()
        user.username = username
        user.password_hash = generate_password_hash(password)
        user.role = role
        user.is_manager = bool(is_manager)
        db.session.add(user)
        db.session.commit()
        # Initialiser les droits (admin a tout, autres rien)
        for page in PAGE_NAMES:
            access = UserPageAccess(user_id=user.id, page_name=page, can_access=(role == 'admin'))
            db.session.add(access)
        db.session.commit()
        return jsonify({'success': True})

@app.route('/api/utilisateurs/<int:user_id>', methods=['DELETE'])
@require_page_access('utilisateurs')
def api_supprimer_utilisateur(user_id):
    user = User.query.get_or_404(user_id)
    if user.username == 'admin':
        return jsonify({'error': 'Impossible de supprimer le compte admin'}), 403
    db.session.delete(user)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/utilisateurs/<int:user_id>/droits', methods=['PUT'])
@require_page_access('utilisateurs')
def api_modifier_droits(user_id):
    user = User.query.get_or_404(user_id)
    if user.role == 'admin':
        return jsonify({'error': 'Impossible de modifier les droits d\'un admin'}), 403
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Données JSON manquantes'}), 400
    
    print(f"DEBUG: Modification des droits pour l'utilisateur {user_id}: {data}")
    
    for page, can_access in data.items():
        access = UserPageAccess.query.filter_by(user_id=user.id, page_name=page).first()
        if access:
            access.can_access = bool(can_access)
            print(f"DEBUG: Mise à jour du droit {page} = {can_access}")
        else:
            # Créer l'enregistrement s'il n'existe pas
            access = UserPageAccess(user_id=user.id, page_name=page, can_access=bool(can_access))
            db.session.add(access)
            print(f"DEBUG: Création du droit {page} = {can_access}")
    
    try:
        db.session.commit()
        print(f"DEBUG: Commit réussi pour l'utilisateur {user_id}")
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        print(f"DEBUG: Erreur lors du commit: {e}")
        return jsonify({'error': f'Erreur lors de la sauvegarde: {str(e)}'}), 500

@app.route('/api/utilisateurs/<int:user_id>', methods=['PUT'])
@require_page_access('utilisateurs')
def api_modifier_utilisateur(user_id):
    user = User.query.get_or_404(user_id)
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Données JSON manquantes'}), 400
    username = data.get('username')
    role = data.get('role')
    password = data.get('password')
    is_manager = data.get('is_manager')
    if username:
        # Vérifier unicité si changement
        if username != user.username and User.query.filter_by(username=username).first():
            return jsonify({'error': 'Nom d\'utilisateur déjà pris'}), 400
        user.username = username
    if role and user.role != 'admin':
        user.role = role
    if password:
        user.password_hash = generate_password_hash(password)
    if 'is_manager' in data:
        user.is_manager = bool(is_manager)
    db.session.commit()
    return jsonify({'success': True})

# Routes pour la gestion du personnel
@app.route('/perso')
@require_page_access('perso')
def perso():
    # Les chefs d'équipe ont la même interface que les opérateurs pour le moment
    is_manager = (current_user.is_manager or current_user.role == 'admin') and current_user.role != 'chef_equipe'
    return render_template('perso.html', is_manager=is_manager)

# API - Liste du personnel (manager voit tout, personnel voit seulement son profil)
@app.route('/api/perso/personnel', methods=['GET'])
@require_page_access('perso')
def api_liste_personnel():
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    if is_manager:
        # Manager voit tout le personnel - précharger la relation user pour éviter N+1
        personnel_list = Personnel.query.options(joinedload(Personnel.user)).all()
    else:
        # Personnel voit seulement son profil
        personnel = Personnel.query.options(joinedload(Personnel.user)).filter_by(user_id=current_user.id).first()
        personnel_list = [personnel] if personnel else []
    
    data = []
    for p in personnel_list:
        data.append({
            'id': p.id,
            'user_id': p.user_id,
            'nom': p.nom,
            'prenom': p.prenom,
            'email': p.email,
            'telephone': p.telephone,
            'date_embauche': p.date_embauche.isoformat() if p.date_embauche else None,
            'poste': p.poste,
            'societe': p.societe,
            'site_id': p.site_id,
            'username': p.user.username if p.user else None
        })
    return jsonify(data)

# API - Créer un membre du personnel (manager seulement)
@app.route('/api/perso/personnel', methods=['POST'])
@require_page_access('perso')
def api_creer_personnel():
    is_manager = current_user.is_manager or current_user.role == 'admin'
    if not is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Données JSON manquantes'}), 400
    
    user_id = data.get('user_id')
    if not user_id:
        return jsonify({'error': 'user_id manquant'}), 400
    
    # Vérifier que l'utilisateur existe et n'a pas déjà un profil personnel
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'Utilisateur introuvable'}), 404
    
    if Personnel.query.filter_by(user_id=user_id).first():
        return jsonify({'error': 'Ce utilisateur a déjà un profil personnel'}), 400
    
    personnel = Personnel(
        user_id=user_id,
        nom=data.get('nom', ''),
        prenom=data.get('prenom', ''),
        email=data.get('email'),
        telephone=data.get('telephone'),
        date_embauche=datetime.strptime(data['date_embauche'], '%Y-%m-%d').date() if data.get('date_embauche') else None,
        poste=data.get('poste'),
        societe=data.get('societe'),
        site_id=int(data.get('site_id')) if data.get('site_id') else None
    )
    db.session.add(personnel)
    db.session.commit()
    return jsonify({'success': True, 'id': personnel.id})

# API - Modifier un membre du personnel
@app.route('/api/perso/personnel/<int:personnel_id>', methods=['PUT'])
@require_page_access('perso')
def api_modifier_personnel(personnel_id):
    personnel = Personnel.query.get_or_404(personnel_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    # Vérifier les droits: manager peut modifier tout, personnel peut modifier seulement son profil
    if not is_manager and personnel.user_id != current_user.id:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Données JSON manquantes'}), 400
    
    if 'nom' in data:
        personnel.nom = data['nom']
    if 'prenom' in data:
        personnel.prenom = data['prenom']
    if 'email' in data:
        personnel.email = data['email']
    if 'telephone' in data:
        personnel.telephone = data['telephone']
    if 'date_embauche' in data:
        personnel.date_embauche = datetime.strptime(data['date_embauche'], '%Y-%m-%d').date() if data['date_embauche'] else None
    if 'poste' in data:
        personnel.poste = data['poste']
    if 'societe' in data:
        personnel.societe = data['societe']
    if 'site_id' in data:
        site_id_value = data.get('site_id')
        if site_id_value:
            try:
                personnel.site_id = int(site_id_value)
            except (ValueError, TypeError):
                personnel.site_id = None
        else:
            personnel.site_id = None
    
    db.session.commit()
    return jsonify({'success': True})

# API - Supprimer un membre du personnel (manager seulement)
@app.route('/api/perso/personnel/<int:personnel_id>', methods=['DELETE'])
@require_page_access('perso')
def api_supprimer_personnel(personnel_id):
    is_manager = current_user.is_manager or current_user.role == 'admin'
    if not is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    personnel = Personnel.query.get_or_404(personnel_id)
    db.session.delete(personnel)
    db.session.commit()
    return jsonify({'success': True})

# API - Gérer les jours travaillés
@app.route('/api/perso/personnel/<int:personnel_id>/jours-travailles', methods=['GET', 'POST', 'PUT'])
@require_page_access('perso')
def api_jours_travailles(personnel_id):
    personnel = Personnel.query.get_or_404(personnel_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    if request.method == 'GET':
        jours = WorkingDays.query.filter_by(personnel_id=personnel_id).all()
        data = [{
            'id': j.id,
            'jour_semaine': j.jour_semaine,
            'type_journee': j.type_journee
        } for j in jours]
        return jsonify(data)
    
    if not is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    if request.method == 'POST' or request.method == 'PUT':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Données JSON manquantes'}), 400
        
        # Supprimer les anciens jours travaillés
        WorkingDays.query.filter_by(personnel_id=personnel_id).delete()
        
        # Créer les nouveaux jours travaillés
        for jour_data in data.get('jours', []):
            jour = WorkingDays(
                personnel_id=personnel_id,
                jour_semaine=jour_data['jour_semaine'],
                type_journee=jour_data['type_journee']
            )
            db.session.add(jour)
        
        db.session.commit()
        # Invalider le cache des jours travaillés pour ce personnel
        _clear_jours_travailles_cache(personnel_id)
        return jsonify({'success': True})

# API - Demandes de congé
@app.route('/api/perso/conges', methods=['GET', 'POST'])
@require_page_access('perso')
def api_conges():
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    if request.method == 'GET':
        if is_manager:
            # Manager voit toutes les demandes - précharger les relations pour éviter N+1
            demandes = LeaveRequest.query.options(joinedload(LeaveRequest.personnel)).order_by(LeaveRequest.created_at.desc()).all()
        else:
            # Personnel voit seulement ses demandes
            personnel = Personnel.query.filter_by(user_id=current_user.id).first()
            if not personnel:
                return jsonify([])
            demandes = LeaveRequest.query.options(joinedload(LeaveRequest.personnel)).filter_by(personnel_id=personnel.id).order_by(LeaveRequest.created_at.desc()).all()
        
        # Précharger tous les jours travaillés en une seule requête
        personnel_ids = list(set([d.personnel_id for d in demandes]))
        all_jours_travailles = WorkingDays.query.filter(WorkingDays.personnel_id.in_(personnel_ids)).all()
        jours_par_personnel = {}
        for jt in all_jours_travailles:
            if jt.personnel_id not in jours_par_personnel:
                jours_par_personnel[jt.personnel_id] = {}
            jours_par_personnel[jt.personnel_id][jt.jour_semaine] = jt.type_journee
        
        data = []
        for d in demandes:
            # Utiliser les jours travaillés préchargés
            jours_semaine = jours_par_personnel.get(d.personnel_id, {})
            nombre_jours = calculer_jours_travailles(d.personnel_id, d.date_debut, d.date_fin, jours_semaine)
            data.append({
                'id': d.id,
                'personnel_id': d.personnel_id,
                'personnel_nom': f"{d.personnel.prenom} {d.personnel.nom}",
                'personnel_prenom': d.personnel.prenom,
                'personnel_nom_complet': d.personnel.nom,
                'personnel_email': d.personnel.email,
                'personnel_site_id': d.personnel.site_id,  # 1=SMP, 2=LPZ
                'date_debut': d.date_debut.isoformat(),
                'date_fin': d.date_fin.isoformat(),
                'type_conge': d.type_conge,
                'statut': d.statut,
                'commentaire': d.commentaire,
                'nombre_jours': nombre_jours,
                'created_at': d.created_at.isoformat()
            })
        return jsonify(data)
    
    elif request.method == 'POST':
        # Créer une demande de congé
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Données JSON manquantes'}), 400
        
        personnel = Personnel.query.filter_by(user_id=current_user.id).first()
        if not personnel:
            return jsonify({'error': 'Profil personnel introuvable'}), 404
        
        demande = LeaveRequest(
            personnel_id=personnel.id,
            date_debut=datetime.strptime(data['date_debut'], '%Y-%m-%d').date(),
            date_fin=datetime.strptime(data['date_fin'], '%Y-%m-%d').date(),
            type_conge=data.get('type_conge', 'conge_paye'),
            commentaire=data.get('commentaire')
        )
        db.session.add(demande)
        db.session.commit()
        return jsonify({'success': True, 'id': demande.id})

# Cache pour les jours travaillés (évite les requêtes répétées)
_jours_travailles_cache = {}

def _get_jours_travailles_dict(personnel_id):
    """Récupère le dictionnaire des jours travaillés avec cache"""
    if personnel_id not in _jours_travailles_cache:
        jours_travailles = WorkingDays.query.filter_by(personnel_id=personnel_id).all()
        _jours_travailles_cache[personnel_id] = {j.jour_semaine: j.type_journee for j in jours_travailles}
    return _jours_travailles_cache[personnel_id]

def _clear_jours_travailles_cache(personnel_id=None):
    """Efface le cache des jours travaillés"""
    if personnel_id:
        _jours_travailles_cache.pop(personnel_id, None)
    else:
        _jours_travailles_cache.clear()

# Fonction pour calculer le nombre de jours travaillés entre deux dates
def calculer_jours_travailles(personnel_id, date_debut, date_fin, jours_semaine=None):
    """Calcule le nombre de jours travaillés entre deux dates en tenant compte des jours travaillés définis"""
    # Utiliser le cache si jours_semaine n'est pas fourni
    if jours_semaine is None:
        jours_semaine = _get_jours_travailles_dict(personnel_id)
    
    # Si aucun jour travaillé n'est défini, compter tous les jours (comportement par défaut)
    if not jours_semaine:
        return (date_fin - date_debut).days + 1
    
    # Compter uniquement les jours travaillés
    nombre_jours = 0
    current_date = date_debut
    while current_date <= date_fin:
        jour_semaine = current_date.weekday()  # 0=lundi, 6=dimanche
        # Si ce jour de la semaine est un jour travaillé, l'ajouter au compte
        if jour_semaine in jours_semaine:
            nombre_jours += 1
        current_date += timedelta(days=1)
    
    return nombre_jours

# Fonction pour générer le PDF de congé en utilisant le template vierge
def generer_pdf_conge(demande):
    """Génère un PDF de formulaire d'autorisation d'absence en remplissant le template vierge"""
    try:
        # Vérifier si le template existe (essayer plusieurs chemins et noms possibles)
        template_paths = [
            TEMPLATE_PDF_PATH,
            os.path.join('static', 'templates', 'formulaire_absence_vierge.pdf'),
            os.path.join('static', 'templates', 'formulaire-absence-vierge.pdf'),  # Avec tirets
            os.path.join(os.path.dirname(__file__), 'static', 'templates', 'formulaire_absence_vierge.pdf'),
            os.path.join(os.path.dirname(__file__), 'static', 'templates', 'formulaire-absence-vierge.pdf'),  # Avec tirets
            os.path.abspath(os.path.join('static', 'templates', 'formulaire_absence_vierge.pdf')),
            os.path.abspath(os.path.join('static', 'templates', 'formulaire-absence-vierge.pdf'))  # Avec tirets
        ]
        
        template_path = None
        for path in template_paths:
            if os.path.exists(path):
                template_path = path
                break
        
        if not template_path:
            error_msg = f"⚠️ Template PDF non trouvé!\n"
            error_msg += f"   Chemins testés:\n"
            for path in template_paths:
                abs_path = os.path.abspath(path) if path else "None"
                exists = os.path.exists(path) if path else False
                error_msg += f"   - {abs_path} (existe: {exists})\n"
            error_msg += f"\n   Veuillez placer le fichier 'formulaire_absence_vierge.pdf' dans le dossier:\n"
            error_msg += f"   {os.path.abspath('static/templates')}\n"
            print(error_msg)
            # Afficher aussi les fichiers présents dans le dossier templates
            templates_dir = os.path.join('static', 'templates')
            if os.path.exists(templates_dir):
                files = os.listdir(templates_dir)
                if files:
                    print(f"   Fichiers trouvés dans static/templates/: {', '.join(files)}")
                else:
                    print(f"   Le dossier static/templates/ existe mais est vide.")
            return None
        
        # Lire le template PDF
        template_reader = PdfReader(open(template_path, 'rb'))
        template_page = template_reader.pages[0]
        
        # Créer un overlay avec reportlab pour écrire les données
        overlay_buffer = io.BytesIO()
        overlay = Canvas(overlay_buffer, pagesize=A4)
        width, height = A4
        
        # Informations du personnel
        personnel = demande.personnel
        
        # Conversion : reportlab utilise Y depuis le bas, l'utilisateur donne Y depuis le haut
        # Hauteur A4 = 29.7cm, donc Y_bas = 29.7 - Y_haut
        height_cm = 29.7
        
        # Date de la demande : X=6.5cm, Y=3.9cm - 4mm = 3.5cm (depuis le haut) - pas de changement
        date_demande = demande.created_at.strftime('%d/%m/%Y') if demande.created_at else datetime.now().strftime('%d/%m/%Y')
        overlay.setFont("Helvetica", 10)
        overlay.drawString(6.5*cm, (height_cm - 3.5)*cm, date_demande)
        
        # NOM : X=4.5cm, Y=4.6cm + 2mm = 4.8cm (depuis le haut)
        overlay.setFont("Helvetica", 11)
        overlay.drawString(4.5*cm, (height_cm - 4.8)*cm, personnel.nom.upper())
        
        # Société d'appartenance : X=7.5cm (3cm à droite du NOM), Y=5.3cm
        if personnel.societe:
            overlay.setFont("Helvetica", 10)
            overlay.drawString(7.5*cm, (height_cm - 5.3)*cm, personnel.societe)
        
        # Prénom : X=14cm, Y=4.6cm + 2mm = 4.8cm (depuis le haut)
        overlay.setFont("Helvetica", 11)
        overlay.drawString(14*cm, (height_cm - 4.8)*cm, personnel.prenom)
        
        # Direction "MATERIEL" : X=5cm, Y=5.6cm + 2mm = 5.8cm (depuis le haut)
        overlay.setFont("Helvetica", 11)
        overlay.drawString(5*cm, (height_cm - 5.8)*cm, "MATERIEL")
        
        # Zone de travaux "STE" : X=6.5cm, Y=6.2cm + 2mm = 6.4cm (depuis le haut)
        overlay.drawString(6.5*cm, (height_cm - 6.4)*cm, "STE")
        
        # Mapper les types de congé vers les colonnes du formulaire
        type_mapping = {
            'conge_paye': 'CP',
            'rtt': 'RTT',
            'conge_sans_solde': 'Congé sans solde',
            'conge_paternite': 'Congé paternité/maternité',
            'congé_autorise_paye': 'Absence autorisée payée',
            'congé_autorise_non_paye': 'Absence autorisée non-payée'
        }
        type_label = type_mapping.get(demande.type_conge, demande.type_conge)
        
        # Calculer le nombre de jours travaillés (en excluant les jours non travaillés)
        # Utiliser le cache pour éviter les requêtes répétées
        jours_semaine = _get_jours_travailles_dict(demande.personnel_id)
        nombre_jours = calculer_jours_travailles(demande.personnel_id, demande.date_debut, demande.date_fin, jours_semaine)
        
        # Dates formatées
        date_debut_str = demande.date_debut.strftime('%d/%m/%Y')
        date_fin_str = demande.date_fin.strftime('%d/%m/%Y')
        
        # Position des dates et du nombre de jours :
        # - Congé payé et congé sans solde : position actuelle
        # - RTT : décalage de 4 cm vers la gauche pour "Du", "Au" et le nombre de jours
        overlay.setFont("Helvetica", 11)
        if demande.type_conge == 'rtt':
            nb_jours_x = 8*cm    # 12cm - 4cm
            dates_x = 7*cm       # 11cm - 4cm
        else:
            nb_jours_x = 12*cm
            dates_x = 11*cm
        
        # Nb de jours : Y=8cm + 2mm = 8.2cm (depuis le haut)
        overlay.drawString(nb_jours_x, (height_cm - 8.2)*cm, str(nombre_jours))
        
        # Du : Y=8.9cm + 2mm = 9.1cm (depuis le haut)
        overlay.drawString(dates_x, (height_cm - 9.1)*cm, date_debut_str)
        
        # Au : Y=9.4cm + 2mm + 2mm = 9.8cm (depuis le haut)
        overlay.drawString(dates_x, (height_cm - 9.8)*cm, date_fin_str)
        
        # Commentaire (si présent, dans la zone commentaire)
        if demande.commentaire:
            overlay.setFont("Helvetica", 9)
            commentaire_y = height - 15.5*cm
            commentaire_lines = demande.commentaire.split('\n')
            for i, line in enumerate(commentaire_lines[:5]):  # Max 5 lignes
                overlay.drawString(2*cm, commentaire_y - (i * 0.5*cm), line[:100])
        
        # Date signature de l'intéressé : X=6.5cm, Y=19.6cm + 1mm = 19.7cm (depuis le haut) - date de demande
        date_signature_interesse = demande.created_at.strftime('%d/%m/%Y') if demande.created_at else datetime.now().strftime('%d/%m/%Y')
        overlay.setFont("Helvetica", 10)
        overlay.drawString(6.5*cm, (height_cm - 19.7)*cm, date_signature_interesse)
        
        # Date signature du supérieur hiérarchique : X=12cm, Y=19.6cm + 1mm = 19.7cm (depuis le haut) - date d'acceptation
        date_signature_superieur = datetime.now().strftime('%d/%m/%Y')
        overlay.drawString(12*cm, (height_cm - 19.7)*cm, date_signature_superieur)
        
        # Fonction helper pour charger et afficher une signature
        def ajouter_signature(user_id, x_cm, role_description=""):
            """Ajoute une signature à la position X donnée, à la hauteur Y=21.2cm"""
            print(f"[PDF] Recherche de la signature pour user_id={user_id} ({role_description})")
            signature = ManagerSignature.query.filter_by(user_id=user_id).first()
            if not signature:
                print(f"[PDF] ⚠️ Aucune signature trouvée en base pour user_id={user_id} ({role_description})")
                return False
            if not os.path.exists(signature.signature_path):
                print(f"[PDF] ⚠️ Fichier de signature introuvable: {signature.signature_path} pour user_id={user_id} ({role_description})")
                return False
            try:
                print(f"[PDF] ✓ Signature trouvée pour user_id={user_id} ({role_description}): {signature.signature_path}")
                # Charger et redimensionner la signature
                img = Image.open(signature.signature_path)
                # Redimensionner pour qu'elle fasse environ 3cm de large (réduite pour ne pas chevaucher avec les dates)
                img_width_cm = 3
                img_width_px = int(img_width_cm * cm)
                ratio = img_width_px / img.width
                img_height_px = int(img.height * ratio)
                img = img.resize((img_width_px, img_height_px), Image.Resampling.LANCZOS)
                
                # Sauvegarder temporairement
                temp_img_path = os.path.join(tempfile.gettempdir(), f"signature_{user_id}_{datetime.now().timestamp()}.png")
                img.save(temp_img_path)
                
                # Position de la signature : Y=21.2cm (depuis le haut) - montée de 0.5cm
                signature_x = x_cm * cm
                signature_y = (height_cm - 21.2) * cm
                overlay.drawImage(temp_img_path, signature_x, signature_y, width=img_width_px, height=img_height_px)
                print(f"[PDF] ✓ Signature ajoutée à X={x_cm}cm, Y=21.2cm pour user_id={user_id} ({role_description})")
                
                # Supprimer le fichier temporaire
                os.remove(temp_img_path)
                return True
            except Exception as e:
                print(f"[PDF] ❌ Erreur lors de l'ajout de la signature (user_id={user_id}, {role_description}): {e}")
                import traceback
                traceback.print_exc()
                return False
        
        # Ajouter la signature de l'opérateur/chef d'équipe (demandeur) : X=5cm (3cm vers la gauche par rapport à 8cm)
        print(f"[PDF] Génération PDF pour demande {demande.id}, personnel_id={personnel.id}, personnel.user_id={personnel.user_id}")
        if personnel.user_id:
            ajouter_signature(personnel.user_id, 5, "demandeur (opérateur/chef d'équipe)")
        else:
            print(f"[PDF] ⚠️ personnel.user_id est None pour personnel_id={personnel.id}")
        
        # Ajouter la signature du manager (qui accepte) : X=12cm
        print(f"[PDF] Manager qui accepte: current_user.id={current_user.id}, current_user.username={current_user.username}")
        ajouter_signature(current_user.id, 12, "manager (qui accepte)")
        
        # Finaliser l'overlay
        overlay.save()
        overlay_buffer.seek(0)
        
        # Fusionner le template et l'overlay
        template_writer = PdfWriter()
        overlay_reader = PdfReader(overlay_buffer)
        overlay_page = overlay_reader.pages[0]
        
        # Fusionner les pages
        template_page.merge_page(overlay_page)
        template_writer.add_page(template_page)
        
        # Créer le PDF final
        output_buffer = io.BytesIO()
        template_writer.write(output_buffer)
        output_buffer.seek(0)
        
        return output_buffer
        
    except Exception as e:
        print(f"Erreur lors de la génération du PDF: {e}")
        import traceback
        traceback.print_exc()
        return None

# API - Accepter/Refuser une demande de congé (manager seulement)
@app.route('/api/perso/conges/<int:demande_id>', methods=['PUT'])
@require_page_access('perso')
def api_modifier_conge(demande_id):
    is_manager = current_user.is_manager or current_user.role == 'admin'
    if not is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    demande = LeaveRequest.query.get_or_404(demande_id)
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Données JSON manquantes'}), 400
    
    if 'statut' in data:
        old_statut = demande.statut
        demande.statut = data['statut']
        
        # Générer le PDF automatiquement si le congé est accepté
        if data['statut'] == 'accepte' and old_statut != 'accepte':
            pdf_buffer = generer_pdf_conge(demande)
            if pdf_buffer:
                try:
                    # Créer le dossier pour les PDFs de congé s'il n'existe pas
                    pdf_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'conges', str(demande_id))
                    os.makedirs(pdf_dir, exist_ok=True)
                    
                    # Nom du fichier PDF
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    pdf_filename = f"formulaire_conge_{demande.id}_{timestamp}.pdf"
                    pdf_path = os.path.join(pdf_dir, pdf_filename)
                    
                    # Sauvegarder le PDF
                    with open(pdf_path, 'wb') as f:
                        f.write(pdf_buffer.read())
                    
                    # Enregistrer le document en base
                    pdf_doc = LeaveRequestDocument(
                        leave_request_id=demande.id,
                        nom_fichier=f"Formulaire d'autorisation d'absence - {demande.personnel.nom} {demande.personnel.prenom}.pdf",
                        chemin_fichier=pdf_path
                    )
                    db.session.add(pdf_doc)
                    
                    # Ajouter le document à la liste de documents du demandeur
                    personnel_doc = PersonnelDocument(
                        personnel_id=demande.personnel_id,
                        nom_fichier=f"Formulaire d'autorisation d'absence - {demande.personnel.nom} {demande.personnel.prenom}.pdf",
                        chemin_fichier=pdf_path,
                        type_document='autorisation_absence',
                        description=f"Formulaire d'autorisation d'absence du {demande.date_debut.strftime('%d/%m/%Y')} au {demande.date_fin.strftime('%d/%m/%Y')}",
                        uploaded_by=current_user.id
                    )
                    db.session.add(personnel_doc)
                except Exception as e:
                    print(f"Erreur lors de la sauvegarde du PDF: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f"⚠️ Échec de la génération du PDF pour le congé {demande_id}")
                print(f"   Vérifiez que le template existe à: {TEMPLATE_PDF_PATH}")
                print(f"   Chemin absolu: {os.path.abspath(TEMPLATE_PDF_PATH)}")
    
    db.session.commit()
    return jsonify({'success': True})

# API - Supprimer un congé (manager seulement)
@app.route('/api/perso/conges/<int:demande_id>', methods=['DELETE'])
@require_page_access('perso')
def api_supprimer_conge(demande_id):
    is_manager = current_user.is_manager or current_user.role == 'admin'
    if not is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    demande = LeaveRequest.query.get_or_404(demande_id)
    
    try:
        # Récupérer tous les documents associés au congé
        documents = LeaveRequestDocument.query.filter_by(leave_request_id=demande_id).all()
        
        # Pour chaque document, supprimer aussi le document PersonnelDocument correspondant
        # (celui qui a été créé lors de la validation)
        for doc in documents:
            # Chercher le document PersonnelDocument qui a le même chemin_fichier
            personnel_doc = PersonnelDocument.query.filter_by(
                personnel_id=demande.personnel_id,
                chemin_fichier=doc.chemin_fichier,
                type_document='autorisation_absence'
            ).first()
            
            if personnel_doc:
                # Supprimer le fichier physique s'il existe
                if os.path.exists(personnel_doc.chemin_fichier):
                    try:
                        os.remove(personnel_doc.chemin_fichier)
                    except Exception as e:
                        print(f"Erreur lors de la suppression du fichier {personnel_doc.chemin_fichier}: {e}")
                
                # Supprimer le document PersonnelDocument
                db.session.delete(personnel_doc)
            
            # Supprimer aussi le fichier physique du LeaveRequestDocument
            if os.path.exists(doc.chemin_fichier):
                try:
                    os.remove(doc.chemin_fichier)
                except Exception as e:
                    print(f"Erreur lors de la suppression du fichier {doc.chemin_fichier}: {e}")
        
        # Supprimer le dossier du congé s'il est vide
        pdf_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'conges', str(demande_id))
        if os.path.exists(pdf_dir):
            try:
                # Essayer de supprimer le dossier s'il est vide
                if not os.listdir(pdf_dir):
                    os.rmdir(pdf_dir)
            except Exception as e:
                print(f"Erreur lors de la suppression du dossier {pdf_dir}: {e}")
        
        # Supprimer le congé (les LeaveRequestDocument seront supprimés automatiquement par cascade)
        db.session.delete(demande)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        print(f"Erreur lors de la suppression du congé: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erreur lors de la suppression: {str(e)}'}), 500

# API - Uploader le template PDF (admin seulement)
@app.route('/api/perso/manager/template-pdf', methods=['POST'])
@require_page_access('perso')
def api_upload_template_pdf():
    if current_user.role != 'admin':
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400
    
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Le fichier doit être un PDF'}), 400
    
    try:
        # Sauvegarder le template
        file.save(TEMPLATE_PDF_PATH)
        return jsonify({'success': True, 'message': 'Template PDF enregistré avec succès'})
    except Exception as e:
        print(f"Erreur lors de l'upload du template: {e}")
        return jsonify({'error': f'Erreur lors de l\'upload: {str(e)}'}), 500

# API - Uploader la signature (opérateurs, chefs d'équipe et managers)
@app.route('/api/perso/manager/signature', methods=['POST'])
@require_page_access('perso')
def api_upload_signature():
    # Permettre à tous les utilisateurs (opérateurs, chefs d'équipe, managers) d'uploader leur signature
    # Vérifier que l'utilisateur a un rôle valide
    if current_user.role not in ['operateur', 'chef_equipe', 'admin'] and not current_user.is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400
    
    # Vérifier que c'est une image
    if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
        return jsonify({'error': 'Le fichier doit être une image'}), 400
    
    try:
        print(f"[UPLOAD SIGNATURE] Upload pour user_id={current_user.id}, username={current_user.username}, role={current_user.role}")
        # Sauvegarder le fichier
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"signature_{current_user.id}_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'signatures')
        os.makedirs(filepath, exist_ok=True)
        full_path = os.path.join(filepath, filename)
        file.save(full_path)
        print(f"[UPLOAD SIGNATURE] Fichier sauvegardé: {full_path}")
        
        # Enregistrer ou mettre à jour la signature
        signature = ManagerSignature.query.filter_by(user_id=current_user.id).first()
        if signature:
            print(f"[UPLOAD SIGNATURE] Mise à jour signature existante (id={signature.id})")
            # Supprimer l'ancienne signature
            if os.path.exists(signature.signature_path):
                os.remove(signature.signature_path)
                print(f"[UPLOAD SIGNATURE] Ancienne signature supprimée: {signature.signature_path}")
            signature.signature_path = full_path
            signature.updated_at = datetime.utcnow()
        else:
            print(f"[UPLOAD SIGNATURE] Création nouvelle signature pour user_id={current_user.id}")
            signature = ManagerSignature(
                user_id=current_user.id,
                signature_path=full_path
            )
            db.session.add(signature)
        
        db.session.commit()
        print(f"[UPLOAD SIGNATURE] ✓ Signature enregistrée avec succès pour user_id={current_user.id}")
        return jsonify({'success': True, 'message': 'Signature enregistrée avec succès'})
    except Exception as e:
        print(f"Erreur lors de l'upload de la signature: {e}")
        return jsonify({'error': f'Erreur lors de l\'upload: {str(e)}'}), 500

# API - Modifier le mot de passe de l'utilisateur connecté
@app.route('/api/perso/mon-compte/password', methods=['PUT'])
@require_page_access('perso')
def api_modifier_mon_password():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Données JSON manquantes'}), 400
    
    ancien_password = data.get('ancien_password')
    nouveau_password = data.get('nouveau_password')
    
    if not ancien_password or not nouveau_password:
        return jsonify({'error': 'Ancien et nouveau mot de passe requis'}), 400
    
    # Vérifier l'ancien mot de passe
    if not check_password_hash(current_user.password_hash, ancien_password):
        return jsonify({'error': 'Ancien mot de passe incorrect'}), 400
    
    # Vérifier que le nouveau mot de passe est différent
    if check_password_hash(current_user.password_hash, nouveau_password):
        return jsonify({'error': 'Le nouveau mot de passe doit être différent de l\'ancien'}), 400
    
    # Mettre à jour le mot de passe
    current_user.password_hash = generate_password_hash(nouveau_password)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Mot de passe modifié avec succès'})

# API - Modifier mes informations personnelles (email, téléphone)
@app.route('/api/perso/mon-compte/infos', methods=['PUT'])
@require_page_access('perso')
def api_modifier_mes_infos():
    # Récupérer le personnel associé à l'utilisateur connecté
    personnel = Personnel.query.filter_by(user_id=current_user.id).first()
    if not personnel:
        return jsonify({'error': 'Profil personnel non trouvé'}), 404
    
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Données JSON manquantes'}), 400
    
    if 'email' in data:
        personnel.email = data['email']
    if 'telephone' in data:
        personnel.telephone = data['telephone']
    
    db.session.commit()
    return jsonify({'success': True, 'message': 'Informations modifiées avec succès'})

# API - Admin: Importer la signature d'un personnel
@app.route('/api/perso/admin/personnel/<int:personnel_id>/signature', methods=['POST'])
@require_page_access('perso')
def api_admin_upload_signature_personnel(personnel_id):
    # Vérifier que l'utilisateur est admin
    if current_user.role != 'admin':
        return jsonify({'error': 'Accès non autorisé. Admin seulement.'}), 403
    
    personnel = Personnel.query.get_or_404(personnel_id)
    if not personnel.user_id:
        return jsonify({'error': 'Ce personnel n\'a pas d\'utilisateur associé'}), 400
    
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400
    
    # Vérifier que c'est une image
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
        return jsonify({'error': 'Le fichier doit être une image'}), 400
    
    try:
        # Sauvegarder le fichier
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"signature_{personnel.user_id}_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'signatures')
        os.makedirs(filepath, exist_ok=True)
        full_path = os.path.join(filepath, filename)
        file.save(full_path)
        
        # Enregistrer ou mettre à jour la signature
        signature = ManagerSignature.query.filter_by(user_id=personnel.user_id).first()
        if signature:
            # Supprimer l'ancienne signature
            if os.path.exists(signature.signature_path):
                os.remove(signature.signature_path)
            signature.signature_path = full_path
            signature.updated_at = datetime.utcnow()
        else:
            signature = ManagerSignature(
                user_id=personnel.user_id,
                signature_path=full_path
            )
            db.session.add(signature)
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'Signature importée avec succès pour {personnel.nom} {personnel.prenom}'})
    except Exception as e:
        print(f"Erreur lors de l'upload de la signature: {e}")
        return jsonify({'error': f'Erreur lors de l\'upload: {str(e)}'}), 500

# API - Télécharger un PDF de congé
@app.route('/api/perso/conges/<int:demande_id>/pdf/<int:pdf_id>')
@require_page_access('perso')
def api_download_pdf_conge(demande_id, pdf_id):
    demande = LeaveRequest.query.get_or_404(demande_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    # Vérifier les droits
    if not is_manager and demande.personnel.user_id != current_user.id:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    pdf_doc = LeaveRequestDocument.query.filter_by(id=pdf_id, leave_request_id=demande_id).first_or_404()
    
    # Supporter les anciens chemins relatifs (ex: "uploads/...") et les chemins absolus
    pdf_path = pdf_doc.chemin_fichier
    if not os.path.isabs(pdf_path):
        pdf_path = os.path.join(app.root_path, pdf_path)
    
    if os.path.exists(pdf_path):
        return send_file(pdf_path, as_attachment=True, download_name=pdf_doc.nom_fichier)
    else:
        return jsonify({'error': 'Fichier introuvable'}), 404

# API - Lister les PDFs d'un congé
@app.route('/api/perso/conges/<int:demande_id>/pdfs')
@require_page_access('perso')
def api_list_pdfs_conge(demande_id):
    demande = LeaveRequest.query.get_or_404(demande_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    # Vérifier les droits
    if not is_manager and demande.personnel.user_id != current_user.id:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    pdfs = LeaveRequestDocument.query.filter_by(leave_request_id=demande_id).order_by(LeaveRequestDocument.created_at.desc()).all()
    data = [{
        'id': p.id,
        'nom_fichier': p.nom_fichier,
        'created_at': p.created_at.isoformat()
    } for p in pdfs]
    return jsonify(data)

# API - Documents du personnel
@app.route('/api/perso/personnel/<int:personnel_id>/documents', methods=['GET', 'POST'])
@require_page_access('perso')
def api_documents(personnel_id):
    personnel = Personnel.query.get_or_404(personnel_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    # Vérifier les droits: manager peut voir/ajouter pour tous, personnel seulement pour lui
    if not is_manager and personnel.user_id != current_user.id:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    if request.method == 'GET':
        documents = PersonnelDocument.query.filter_by(personnel_id=personnel_id).all()
        data = [{
            'id': d.id,
            'nom_fichier': d.nom_fichier,
            'type_document': d.type_document,
            'description': d.description,
            'created_at': d.created_at.isoformat()
        } for d in documents]
        return jsonify(data)
    
    elif request.method == 'POST':
        if not is_manager:
            return jsonify({'error': 'Accès non autorisé'}), 403
        
        if 'file' not in request.files:
            return jsonify({'error': 'Aucun fichier fourni'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Aucun fichier sélectionné'}), 400
        
        # Sauvegarder le fichier
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'personnel', str(personnel_id))
        os.makedirs(filepath, exist_ok=True)
        full_path = os.path.join(filepath, filename)
        file.save(full_path)
        
        document = PersonnelDocument(
            personnel_id=personnel_id,
            nom_fichier=file.filename,
            chemin_fichier=full_path,
            type_document=request.form.get('type_document'),
            description=request.form.get('description'),
            uploaded_by=current_user.id
        )
        db.session.add(document)
        db.session.commit()
        return jsonify({'success': True, 'id': document.id})

# API - Télécharger un document
@app.route('/api/perso/documents/<int:document_id>/download')
@require_page_access('perso')
def api_download_document(document_id):
    document = PersonnelDocument.query.get_or_404(document_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    # Vérifier les droits
    if not is_manager and document.personnel.user_id != current_user.id:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    # Supporter les anciens chemins relatifs (ex: "uploads/...") et les chemins absolus
    file_path = document.chemin_fichier
    if not os.path.isabs(file_path):
        file_path = os.path.join(app.root_path, file_path)
    
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=document.nom_fichier)
    else:
        return jsonify({'error': 'Fichier introuvable'}), 404

# API - Supprimer un document (manager seulement)
@app.route('/api/perso/documents/<int:document_id>', methods=['DELETE'])
@require_page_access('perso')
def api_supprimer_document(document_id):
    is_manager = current_user.is_manager or current_user.role == 'admin'
    if not is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    document = PersonnelDocument.query.get_or_404(document_id)
    # Supporter les anciens chemins relatifs et absolus
    file_path = document.chemin_fichier
    if not os.path.isabs(file_path):
        file_path = os.path.join(app.root_path, file_path)
    if os.path.exists(file_path):
        os.remove(file_path)
    db.session.delete(document)
    db.session.commit()
    return jsonify({'success': True})

# API - Absences (manager seulement pour créer)
@app.route('/api/perso/absences', methods=['GET', 'POST'])
@require_page_access('perso')
def api_absences():
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    if request.method == 'GET':
        type_liste = request.args.get('type', 'avenir')  # 'attente' ou 'avenir'
        
        personnel_id = request.args.get('personnel_id', type=int)
        
        if type_liste == 'attente':
            # Absences en attente de validation créées par les employés (non managers)
            # Gérer le cas où le champ statut n'existe pas encore ou est NULL
            try:
                base_query = Absence.query.join(User, Absence.created_by == User.id).filter(
                    or_(
                        Absence.statut == 'en_attente',
                        Absence.statut == None
                    ),
                    User.is_manager == False,
                    User.role != 'admin'
                )
            except:
                # Si le champ statut n'existe pas, filtrer seulement par créateur
                base_query = Absence.query.join(User, Absence.created_by == User.id).filter(
                    User.is_manager == False,
                    User.role != 'admin'
                )
            
            # Précharger les relations personnel pour éviter N+1
            base_query = base_query.options(joinedload(Absence.personnel))
            
            if personnel_id:
                absences = base_query.filter_by(personnel_id=personnel_id).all()
            elif is_manager:
                # Managers voient toutes les demandes des employés en attente
                absences = base_query.all()
            else:
                # Personnel voit seulement ses absences en attente qu'il a créées
                personnel = Personnel.query.filter_by(user_id=current_user.id).first()
                if not personnel:
                    return jsonify([])
                absences = base_query.filter_by(personnel_id=personnel.id).all()
            
            # Sérialiser les absences en attente
            data = []
            for a in absences:
                try:
                    statut = getattr(a, 'statut', None) or 'en_attente'
                except:
                    statut = 'en_attente'
                
                try:
                    personnel_nom = f"{a.personnel.prenom} {a.personnel.nom}" if a.personnel else 'Inconnu'
                except:
                    personnel_nom = 'Inconnu'
                
                data.append({
                    'id': a.id,
                    'personnel_id': a.personnel_id,
                    'personnel_nom': personnel_nom,
                    'date_debut': a.date_debut.isoformat() if a.date_debut else None,
                    'date_fin': a.date_fin.isoformat() if a.date_fin else None,
                    'type_absence': a.type_absence or '-',
                    'statut': statut,
                    'commentaire': a.commentaire or '-',
                    'created_at': a.created_at.isoformat() if a.created_at else None
                })
            
            return jsonify(data)
        else:
            # Absences à venir (mois en cours et mois suivant)
            # Inclure les absences validées ET les congés acceptés
            today = datetime.now().date()
            # Premier jour du mois suivant
            if today.month == 12:
                premier_jour_mois_suivant = datetime(today.year + 1, 1, 1).date()
            else:
                premier_jour_mois_suivant = datetime(today.year, today.month + 1, 1).date()
            # Dernier jour du mois suivant
            if premier_jour_mois_suivant.month == 12:
                dernier_jour_mois_suivant = datetime(premier_jour_mois_suivant.year + 1, 1, 1).date() - timedelta(days=1)
            else:
                dernier_jour_mois_suivant = datetime(premier_jour_mois_suivant.year, premier_jour_mois_suivant.month + 1, 1).date() - timedelta(days=1)
            
            # Récupérer les absences validées
            try:
                base_query_absences = Absence.query.filter(
                    or_(
                        Absence.statut == 'validee',
                        Absence.statut == None  # Pour les absences créées avant l'ajout du champ statut
                    ),
                    Absence.date_fin >= today,
                    Absence.date_debut <= dernier_jour_mois_suivant
                )
            except Exception as e:
                # Si le champ statut n'existe pas, filtrer seulement par date
                print(f"Champ statut non disponible, filtrage par date uniquement: {e}")
                base_query_absences = Absence.query.filter(
                    Absence.date_fin >= today,
                    Absence.date_debut <= dernier_jour_mois_suivant
                )
            
            # Récupérer les congés acceptés
            base_query_conges = LeaveRequest.query.filter(
                LeaveRequest.statut == 'accepte',
                LeaveRequest.date_fin >= today,
                LeaveRequest.date_debut <= dernier_jour_mois_suivant
            )
            
            # Récupérer les formations à venir
            base_query_formations = Formation.query.filter(
                Formation.date_fin >= today,
                Formation.date_debut <= dernier_jour_mois_suivant
            )
            
            # Précharger les relations pour éviter N+1
            base_query_absences = base_query_absences.options(joinedload(Absence.personnel))
            base_query_conges = base_query_conges.options(joinedload(LeaveRequest.personnel))
            base_query_formations = base_query_formations.options(joinedload(Formation.personnel))
            
            if personnel_id:
                absences = base_query_absences.filter_by(personnel_id=personnel_id).all()
                conges = base_query_conges.filter_by(personnel_id=personnel_id).all()
                formations = base_query_formations.filter_by(personnel_id=personnel_id).all()
            elif is_manager:
                absences = base_query_absences.all()
                conges = base_query_conges.all()
                formations = base_query_formations.all()
            else:
                # Personnel voit seulement ses absences, congés et formations
                personnel = Personnel.query.filter_by(user_id=current_user.id).first()
                if not personnel:
                    return jsonify([])
                absences = base_query_absences.filter_by(personnel_id=personnel.id).all()
                conges = base_query_conges.filter_by(personnel_id=personnel.id).all()
                formations = base_query_formations.filter_by(personnel_id=personnel.id).all()
            
            # Combiner les absences, congés et formations
            data = []
            
            # Ajouter les absences validées
            for a in absences:
                try:
                    statut = getattr(a, 'statut', None) or 'en_attente'
                except:
                    statut = 'en_attente'
                
                data.append({
                    'id': a.id,
                    'personnel_id': a.personnel_id,
                    'personnel_nom': f"{a.personnel.prenom} {a.personnel.nom}",
                    'date_debut': a.date_debut.isoformat(),
                    'date_fin': a.date_fin.isoformat(),
                    'type_absence': a.type_absence,
                    'statut': statut,
                    'commentaire': a.commentaire,
                    'created_at': a.created_at.isoformat() if a.created_at else None,
                    'source': 'absence'
                })
            
            # Ajouter les congés acceptés
            for c in conges:
                data.append({
                    'id': c.id,
                    'personnel_id': c.personnel_id,
                    'personnel_nom': f"{c.personnel.prenom} {c.personnel.nom}",
                    'date_debut': c.date_debut.isoformat(),
                    'date_fin': c.date_fin.isoformat(),
                    'type_absence': f"Congé ({c.type_conge})",
                    'statut': 'validee',
                    'commentaire': c.commentaire,
                    'created_at': c.created_at.isoformat() if c.created_at else None,
                    'source': 'conge'
                })
            
            # Ajouter les formations
            for f in formations:
                try:
                    personnel_nom = f"{f.personnel.prenom} {f.personnel.nom}" if f.personnel else 'Inconnu'
                except:
                    personnel_nom = 'Inconnu'
                
                data.append({
                    'id': f.id,
                    'personnel_id': f.personnel_id,
                    'personnel_nom': personnel_nom,
                    'date_debut': f.date_debut.isoformat(),
                    'date_fin': f.date_fin.isoformat(),
                    'type_absence': f"Formation: {f.nom_formation}",
                    'statut': f.statut or 'prevue',
                    'commentaire': f.description or '-',
                    'created_at': f.created_at.isoformat() if f.created_at else None,
                    'source': 'formation'
                })
            
            # Trier par date de début
            data = sorted(data, key=lambda x: x['date_debut'])
            return jsonify(data)
    
    elif request.method == 'POST':
        if not is_manager:
            return jsonify({'error': 'Accès non autorisé'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Données JSON manquantes'}), 400
        
        # Si c'est un manager qui crée l'absence, elle est automatiquement validée
        statut_absence = data.get('statut', 'validee' if is_manager else 'en_attente')
        
        absence = Absence(
            personnel_id=data['personnel_id'],
            date_debut=datetime.strptime(data['date_debut'], '%Y-%m-%d').date(),
            date_fin=datetime.strptime(data['date_fin'], '%Y-%m-%d').date(),
            type_absence=data['type_absence'],
            statut=statut_absence,
            commentaire=data.get('commentaire'),
            created_by=current_user.id
        )
        db.session.add(absence)
        db.session.commit()
        return jsonify({'success': True, 'id': absence.id})

# API - Planning mensuel
@app.route('/api/perso/planning/<int:personnel_id>')
@require_page_access('perso')
def api_planning(personnel_id):
    personnel = Personnel.query.get_or_404(personnel_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    # Vérifier les droits
    if not is_manager and personnel.user_id != current_user.id:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    mois = request.args.get('mois', type=int)
    annee = request.args.get('annee', type=int)
    
    if not mois or not annee:
        today = datetime.now()
        mois = today.month
        annee = today.year
    
    # Récupérer les jours travaillés
    jours_travailles = WorkingDays.query.filter_by(personnel_id=personnel_id).all()
    jours_semaine = {j.jour_semaine: j.type_journee for j in jours_travailles}
    
    # Générer le planning du mois
    premier_jour = datetime(annee, mois, 1).date()
    dernier_jour = (premier_jour.replace(month=premier_jour.month % 12 + 1, day=1) - timedelta(days=1)) if premier_jour.month < 12 else premier_jour.replace(year=premier_jour.year + 1, month=1, day=1) - timedelta(days=1)
    
    # OPTIMISATION : Précharger toutes les données en une seule requête par type
    all_conges = LeaveRequest.query.filter(
        LeaveRequest.personnel_id == personnel_id,
        LeaveRequest.statut.in_(['accepte', 'en_attente']),
        LeaveRequest.date_debut <= dernier_jour,
        LeaveRequest.date_fin >= premier_jour
    ).all()
    conges_par_date = {}
    for c in all_conges:
        date_conge = max(c.date_debut, premier_jour)
        while date_conge <= min(c.date_fin, dernier_jour):
            if date_conge not in conges_par_date:
                conges_par_date[date_conge] = c
            date_conge += timedelta(days=1)
    
    all_absences = Absence.query.filter(
        Absence.personnel_id == personnel_id,
        Absence.date_debut <= dernier_jour,
        Absence.date_fin >= premier_jour
    ).all()
    absences_par_date = {}
    for a in all_absences:
        date_absence = max(a.date_debut, premier_jour)
        while date_absence <= min(a.date_fin, dernier_jour):
            if date_absence not in absences_par_date:
                absences_par_date[date_absence] = a
            date_absence += timedelta(days=1)
    
    all_formations = Formation.query.filter(
        Formation.personnel_id == personnel_id,
        Formation.date_debut <= dernier_jour,
        Formation.date_fin >= premier_jour
    ).all()
    formations_par_date = {}
    for f in all_formations:
        date_formation = max(f.date_debut, premier_jour)
        while date_formation <= min(f.date_fin, dernier_jour):
            if date_formation not in formations_par_date:
                formations_par_date[date_formation] = f
            date_formation += timedelta(days=1)
    
    planning = []
    current_date = premier_jour
    
    while current_date <= dernier_jour:
        jour_semaine = current_date.weekday()  # 0=lundi, 6=dimanche
        
        # Déterminer le type de journée
        type_journee = jours_semaine.get(jour_semaine, None)
        
        # Récupérer depuis les dictionnaires préchargés (pas de requête SQL)
        conge = conges_par_date.get(current_date)
        absence = absences_par_date.get(current_date)
        formation = formations_par_date.get(current_date)
        
        jour_data = {
            'date': current_date.isoformat(),
            'jour_semaine': jour_semaine,
            'type_journee': type_journee,
            'conge': {
                'type': conge.type_conge,
                'statut': conge.statut
            } if conge else None,
            'absence': {
                'type': absence.type_absence,
                'commentaire': absence.commentaire
            } if absence else None,
            'formation': {
                'nom': formation.nom_formation,
                'statut': formation.statut
            } if formation else None
        }
        planning.append(jour_data)
        current_date += timedelta(days=1)
    
    return jsonify(planning)

# API - Planning global mensuel (tous les personnels)
@app.route('/api/perso/planning-global')
@require_page_access('perso')
def api_planning_global():
    try:
        mois = request.args.get('mois', type=int)
        annee = request.args.get('annee', type=int)
        site_id = request.args.get('site_id', type=int)
        
        if not mois or not annee:
            today = datetime.now()
            mois = today.month
            annee = today.year
        
        # Récupérer tous les personnels
        if site_id:
            personnel_list = Personnel.query.filter_by(site_id=site_id).all()
        else:
            personnel_list = Personnel.query.all()
        
        if not personnel_list:
            return jsonify([])
        
        personnel_ids = [p.id for p in personnel_list]
        
        # Calculer les dates du mois
        premier_jour = datetime(annee, mois, 1).date()
        dernier_jour = (premier_jour.replace(month=premier_jour.month % 12 + 1, day=1) - timedelta(days=1)) if premier_jour.month < 12 else premier_jour.replace(year=premier_jour.year + 1, month=1, day=1) - timedelta(days=1)
        
        # OPTIMISATION : Précharger toutes les données en une seule requête par type
        # 1. Tous les jours travaillés
        all_jours_travailles = WorkingDays.query.filter(WorkingDays.personnel_id.in_(personnel_ids)).all()
        jours_par_personnel = {}
        for jt in all_jours_travailles:
            if jt.personnel_id not in jours_par_personnel:
                jours_par_personnel[jt.personnel_id] = {}
            jours_par_personnel[jt.personnel_id][jt.jour_semaine] = jt.type_journee
        
        # 2. Tous les congés du mois
        all_conges = LeaveRequest.query.filter(
            LeaveRequest.personnel_id.in_(personnel_ids),
            LeaveRequest.statut.in_(['accepte', 'en_attente']),
            LeaveRequest.date_debut <= dernier_jour,
            LeaveRequest.date_fin >= premier_jour
        ).all()
        conges_par_personnel_date = {}
        for c in all_conges:
            if c.personnel_id not in conges_par_personnel_date:
                conges_par_personnel_date[c.personnel_id] = {}
            # Créer une entrée pour chaque jour du congé
            date_conge = c.date_debut
            while date_conge <= c.date_fin:
                if date_conge not in conges_par_personnel_date[c.personnel_id]:
                    conges_par_personnel_date[c.personnel_id][date_conge] = c
                date_conge += timedelta(days=1)
        
        # 3. Toutes les absences du mois
        try:
            all_absences = Absence.query.filter(
                Absence.personnel_id.in_(personnel_ids),
                or_(Absence.statut == 'validee', Absence.statut == None),
                Absence.date_debut <= dernier_jour,
                Absence.date_fin >= premier_jour
            ).all()
        except:
            all_absences = Absence.query.filter(
                Absence.personnel_id.in_(personnel_ids),
                Absence.date_debut <= dernier_jour,
                Absence.date_fin >= premier_jour
            ).all()
        absences_par_personnel_date = {}
        for a in all_absences:
            if a.personnel_id not in absences_par_personnel_date:
                absences_par_personnel_date[a.personnel_id] = {}
            date_absence = a.date_debut
            while date_absence <= a.date_fin:
                if date_absence not in absences_par_personnel_date[a.personnel_id]:
                    absences_par_personnel_date[a.personnel_id][date_absence] = a
                date_absence += timedelta(days=1)
        
        # 4. Toutes les formations du mois
        all_formations = Formation.query.filter(
            Formation.personnel_id.in_(personnel_ids),
            Formation.date_debut <= dernier_jour,
            Formation.date_fin >= premier_jour
        ).all()
        formations_par_personnel_date = {}
        for f in all_formations:
            if f.personnel_id not in formations_par_personnel_date:
                formations_par_personnel_date[f.personnel_id] = {}
            date_formation = f.date_debut
            while date_formation <= f.date_fin:
                if date_formation not in formations_par_personnel_date[f.personnel_id]:
                    formations_par_personnel_date[f.personnel_id][date_formation] = f
                date_formation += timedelta(days=1)
        
        # Générer le planning pour chaque personnel (sans requêtes SQL dans la boucle)
        result = []
        for personnel in personnel_list:
            jours_semaine = jours_par_personnel.get(personnel.id, {})
            conges_personnel = conges_par_personnel_date.get(personnel.id, {})
            absences_personnel = absences_par_personnel_date.get(personnel.id, {})
            formations_personnel = formations_par_personnel_date.get(personnel.id, {})
            
            planning = []
            current_date = premier_jour
            
            while current_date <= dernier_jour:
                jour_semaine = current_date.weekday()  # 0=lundi, 6=dimanche
                type_journee = jours_semaine.get(jour_semaine, None)
                
                # Récupérer depuis les dictionnaires préchargés (pas de requête SQL)
                conge = conges_personnel.get(current_date)
                absence = absences_personnel.get(current_date)
                formation = formations_personnel.get(current_date)
                
                jour_data = {
                    'date': current_date.isoformat(),
                    'jour_semaine': jour_semaine,
                    'type_journee': type_journee,
                    'conge': {
                        'type': conge.type_conge,
                        'statut': conge.statut
                    } if conge else None,
                    'absence': {
                        'type': absence.type_absence,
                        'commentaire': absence.commentaire
                    } if absence else None,
                    'formation': {
                        'nom': formation.nom_formation,
                        'statut': formation.statut
                    } if formation else None
                }
                
                planning.append(jour_data)
                current_date += timedelta(days=1)
            
            result.append({
                'personnel_id': personnel.id,
                'personnel_nom': f"{personnel.prenom} {personnel.nom}",
                'planning': planning
            })
        
        return jsonify(result)
    except Exception as e:
        print(f"Erreur dans api_planning_global: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erreur lors de la génération du planning: {str(e)}'}), 500

# API - Modifier le statut d'une absence (manager seulement)
@app.route('/api/perso/absences/<int:absence_id>', methods=['PUT'])
@require_page_access('perso')
def api_modifier_absence(absence_id):
    is_manager = current_user.is_manager or current_user.role == 'admin'
    if not is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    absence = Absence.query.get_or_404(absence_id)
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Données JSON manquantes'}), 400
    
    if 'statut' in data:
        try:
            absence.statut = data['statut']
        except AttributeError:
            # Le champ statut n'existe pas encore
            return jsonify({'error': 'Le champ statut n\'est pas encore disponible. Veuillez redémarrer l\'application pour appliquer la migration.'}), 400
    
    db.session.commit()
    return jsonify({'success': True})

# API - Supprimer une absence (manager seulement)
@app.route('/api/perso/absences/<int:absence_id>', methods=['DELETE'])
@require_page_access('perso')
def api_supprimer_absence(absence_id):
    is_manager = current_user.is_manager or current_user.role == 'admin'
    if not is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    absence = Absence.query.get_or_404(absence_id)
    db.session.delete(absence)
    db.session.commit()
    return jsonify({'success': True})

# API - Liste des utilisateurs pour créer du personnel (manager seulement)
@app.route('/api/perso/utilisateurs-disponibles')
@require_page_access('perso')
def api_utilisateurs_disponibles():
    is_manager = current_user.is_manager or current_user.role == 'admin'
    if not is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    # Utilisateurs qui n'ont pas encore de profil personnel - optimisé
    personnel_user_ids = db.session.query(Personnel.user_id).filter(Personnel.user_id.isnot(None)).all()
    personnel_user_ids = [uid[0] for uid in personnel_user_ids]  # Extraire les IDs de la liste de tuples
    if personnel_user_ids:
        users = User.query.filter(~User.id.in_(personnel_user_ids)).all()
    else:
        users = User.query.all()
    
    data = [{
        'id': u.id,
        'username': u.username,
        'role': u.role
    } for u in users]
    return jsonify(data)

# API - Formations d'un personnel
@app.route('/api/perso/personnel/<int:personnel_id>/formations', methods=['GET', 'POST'])
@require_page_access('perso')
def api_formations(personnel_id):
    personnel = Personnel.query.get_or_404(personnel_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    
    # Vérifier les droits: manager peut voir/ajouter pour tous, personnel seulement pour lui
    if not is_manager and personnel.user_id != current_user.id:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    if request.method == 'GET':
        formations = Formation.query.filter_by(personnel_id=personnel_id).order_by(Formation.date_debut.desc()).all()
        data = []
        for f in formations:
            # Compter les documents
            nb_docs = len(f.documents)
            data.append({
                'id': f.id,
                'nom_formation': f.nom_formation,
                'type_formation': f.type_formation or 'demande',
                'date_debut': f.date_debut.isoformat(),
                'date_fin': f.date_fin.isoformat(),
                'date_fin_validite': f.date_fin_validite.isoformat() if f.date_fin_validite else None,
                'statut': f.statut,
                'description': f.description,
                'created_at': f.created_at.isoformat(),
                'updated_at': f.updated_at.isoformat(),
                'nb_documents': nb_docs
            })
        return jsonify(data)
    
    elif request.method == 'POST':
        if not is_manager:
            return jsonify({'error': 'Accès non autorisé'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Données manquantes'}), 400
        
        date_fin_validite = None
        if data.get('date_fin_validite'):
            date_fin_validite = datetime.strptime(data.get('date_fin_validite'), '%Y-%m-%d').date()
        
        formation = Formation(
            personnel_id=personnel_id,
            nom_formation=data.get('nom_formation'),
            type_formation=data.get('type_formation', 'demande'),
            date_debut=datetime.strptime(data.get('date_debut'), '%Y-%m-%d').date(),
            date_fin=datetime.strptime(data.get('date_fin'), '%Y-%m-%d').date(),
            date_fin_validite=date_fin_validite,
            statut=data.get('statut', 'prevue'),
            description=data.get('description'),
            created_by=current_user.id
        )
        db.session.add(formation)
        db.session.commit()
        return jsonify({'success': True, 'id': formation.id})

# API - Modifier/Supprimer une formation
@app.route('/api/perso/formations/<int:formation_id>', methods=['PUT', 'DELETE'])
@require_page_access('perso')
def api_formation(formation_id):
    formation = Formation.query.get_or_404(formation_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    personnel = formation.personnel
    
    # Vérifier les droits
    if not is_manager and personnel.user_id != current_user.id:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    if request.method == 'PUT':
        if not is_manager:
            return jsonify({'error': 'Accès non autorisé'}), 403
        
        data = request.get_json()
        if data.get('nom_formation'):
            formation.nom_formation = data.get('nom_formation')
        if data.get('type_formation'):
            formation.type_formation = data.get('type_formation')
        if data.get('date_debut'):
            formation.date_debut = datetime.strptime(data.get('date_debut'), '%Y-%m-%d').date()
        if data.get('date_fin'):
            formation.date_fin = datetime.strptime(data.get('date_fin'), '%Y-%m-%d').date()
        if 'date_fin_validite' in data:
            if data.get('date_fin_validite'):
                formation.date_fin_validite = datetime.strptime(data.get('date_fin_validite'), '%Y-%m-%d').date()
            else:
                formation.date_fin_validite = None
        if data.get('statut'):
            formation.statut = data.get('statut')
        if 'description' in data:
            formation.description = data.get('description')
        
        db.session.commit()
        return jsonify({'success': True})
    
    elif request.method == 'DELETE':
        if not is_manager:
            return jsonify({'error': 'Accès non autorisé'}), 403
        
        # Supprimer les documents associés
        for doc in formation.documents:
            if os.path.exists(doc.chemin_fichier):
                os.remove(doc.chemin_fichier)
        
        db.session.delete(formation)
        db.session.commit()
        return jsonify({'success': True})

# API - Documents d'une formation
@app.route('/api/perso/formations/<int:formation_id>/documents', methods=['GET', 'POST'])
@require_page_access('perso')
def api_formation_documents(formation_id):
    formation = Formation.query.get_or_404(formation_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    personnel = formation.personnel
    
    # Vérifier les droits
    if not is_manager and personnel.user_id != current_user.id:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    if request.method == 'GET':
        documents = FormationDocument.query.filter_by(formation_id=formation_id).all()
        data = [{
            'id': d.id,
            'nom_fichier': d.nom_fichier,
            'type_document': d.type_document,
            'description': d.description,
            'created_at': d.created_at.isoformat()
        } for d in documents]
        return jsonify(data)
    
    elif request.method == 'POST':
        if not is_manager:
            return jsonify({'error': 'Accès non autorisé'}), 403
        
        if 'file' not in request.files:
            return jsonify({'error': 'Aucun fichier fourni'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Aucun fichier sélectionné'}), 400
        
        # Sauvegarder le fichier
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'formations', str(formation_id))
        os.makedirs(filepath, exist_ok=True)
        full_path = os.path.join(filepath, filename)
        file.save(full_path)
        
        document = FormationDocument(
            formation_id=formation_id,
            nom_fichier=file.filename,
            chemin_fichier=full_path,
            type_document=request.form.get('type_document'),
            description=request.form.get('description'),
            uploaded_by=current_user.id
        )
        db.session.add(document)
        db.session.commit()
        return jsonify({'success': True, 'id': document.id})

# API - Télécharger un document de formation
@app.route('/api/perso/formation-documents/<int:document_id>/download')
@require_page_access('perso')
def api_download_formation_document(document_id):
    document = FormationDocument.query.get_or_404(document_id)
    is_manager = current_user.is_manager or current_user.role == 'admin'
    personnel = document.formation.personnel
    
    # Vérifier les droits
    if not is_manager and personnel.user_id != current_user.id:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    if os.path.exists(document.chemin_fichier):
        return send_file(document.chemin_fichier, as_attachment=True, download_name=document.nom_fichier)
    else:
        return jsonify({'error': 'Fichier introuvable'}), 404

# API - Supprimer un document de formation
@app.route('/api/perso/formation-documents/<int:document_id>', methods=['DELETE'])
@require_page_access('perso')
def api_supprimer_formation_document(document_id):
    is_manager = current_user.is_manager or current_user.role == 'admin'
    if not is_manager:
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    document = FormationDocument.query.get_or_404(document_id)
    if os.path.exists(document.chemin_fichier):
        os.remove(document.chemin_fichier)
    db.session.delete(document)
    db.session.commit()
    return jsonify({'success': True})

# Fonction de sauvegarde automatique de la base de données
def backup_database():
    """Sauvegarde automatique de la base de données"""
    try:
        db_path = 'instance/ste_releve.db'
        if os.path.exists(db_path):
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = f'instance/backup_ste_releve_{timestamp}.db'
            shutil.copy2(db_path, backup_path)
            print(f"Sauvegarde automatique créée: {backup_path}")
            
            # Garder seulement les 5 dernières sauvegardes
            backup_dir = 'instance'
            backup_files = [f for f in os.listdir(backup_dir) if f.startswith('backup_ste_releve_') and f.endswith('.db')]
            backup_files.sort(reverse=True)
            
            for old_backup in backup_files[5:]:  # Garder seulement les 5 plus récentes
                try:
                    os.remove(os.path.join(backup_dir, old_backup))
                    print(f"Ancienne sauvegarde supprimée: {old_backup}")
                except Exception as e:
                    print(f"Erreur lors de la suppression de {old_backup}: {e}")
    except Exception as e:
        print(f"Erreur lors de la sauvegarde automatique: {e}")

# Fonction de nettoyage automatique de la base de données
def cleanup_old_data():
    """Nettoie automatiquement les anciennes données pour économiser l'espace"""
    try:
        with app.app_context():
            # Supprimer les photos de plus de 2 ans
            two_years_ago = datetime.now().date() - timedelta(days=730)
            old_photos = PhotoReleve.query.filter(PhotoReleve.date < two_years_ago).all()
            
            for photo in old_photos:
                # Supprimer le fichier physique
                try:
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], photo.fichier_photo)
                    if os.path.exists(file_path):
                        os.remove(file_path)
                except Exception as e:
                    print(f"Erreur suppression fichier {photo.fichier_photo}: {e}")
            
            # Supprimer les enregistrements de la base
            PhotoReleve.query.filter(PhotoReleve.date < two_years_ago).delete()
            
            # Supprimer les relevés de plus de 5 ans
            five_years_ago = datetime.now().date() - timedelta(days=1825)
            Releve.query.filter(Releve.date < five_years_ago).delete()
            
            # Supprimer les réponses de routine de plus de 3 ans
            three_years_ago = datetime.now().date() - timedelta(days=1095)
            ReponseRoutine.query.filter(ReponseRoutine.date_creation < three_years_ago).delete()
            
            # Supprimer les codes magasin de plus de 3 mois
            trois_mois_avant = datetime.now().date() - timedelta(days=90)
            codes_supprimes = CodeMagasin.query.filter(CodeMagasin.date < trois_mois_avant).delete()
            
            db.session.commit()
            
            print(f"Nettoyage automatique effectué : {len(old_photos)} photos supprimées, {codes_supprimes} codes magasin supprimés")
            return True
            
    except Exception as e:
        print(f"Erreur lors du nettoyage automatique: {e}")
        return False

# Fonction pour vérifier l'espace utilisé
def check_database_size():
    """Vérifie la taille de la base de données"""
    try:
        with app.app_context():
            # Compter les enregistrements
            nb_releves = Releve.query.count()
            nb_photos = PhotoReleve.query.count()
            nb_routines = ReponseRoutine.query.count()
            nb_codes = CodeMagasin.query.count()
            
            # Estimation de la taille (approximative)
            estimated_size_mb = (nb_releves * 0.01) + (nb_photos * 2) + (nb_routines * 0.01) + (nb_codes * 0.001)
            
            print(f"📊 Taille estimée de la base : {estimated_size_mb:.2f} MB")
            print(f"   - {nb_releves} relevés")
            print(f"   - {nb_photos} photos")
            print(f"   - {nb_routines} réponses de routine")
            print(f"   - {nb_codes} codes magasin")
            
            # Afficher un avertissement si > 800MB (sans déclencher automatiquement)
            if estimated_size_mb > 800:
                print("⚠️ ATTENTION : Base de données proche de la limite (1GB)")
                print("💡 Utilisez le bouton 'Nettoyage automatique' dans l'interface admin pour nettoyer")
            
            return estimated_size_mb
            
    except Exception as e:
        print(f"Erreur lors de la vérification de la taille: {e}")
        return 0

@app.route('/api/database/status')
@login_required
def api_database_status():
    """Retourne le statut de la base de données"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    try:
        # Compter les enregistrements
        nb_releves = Releve.query.count()
        nb_photos = PhotoReleve.query.count()
        nb_routines = ReponseRoutine.query.count()
        nb_users = User.query.count()
        nb_codes = CodeMagasin.query.count()
        
        # Estimation de la taille
        estimated_size_mb = (nb_releves * 0.01) + (nb_photos * 2) + (nb_routines * 0.01) + (nb_codes * 0.001)
        
        # Déterminer le type de base de données
        db_uri = app.config['SQLALCHEMY_DATABASE_URI']
        if 'postgresql' in db_uri or 'postgres' in db_uri:
            db_type = 'PostgreSQL'
            db_icon = 'database'
            db_color = 'primary'
        else:
            db_type = 'SQLite'
            db_icon = 'hdd'
            db_color = 'secondary'
        
        # Statut de l'espace
        if estimated_size_mb > 900:
            status = 'critical'
            message = 'Base de données presque pleine ! Upgrade recommandé.'
        elif estimated_size_mb > 800:
            status = 'warning'
            message = 'Base de données proche de la limite.'
        else:
            status = 'ok'
            message = 'Espace suffisant.'
        
        return jsonify({
            'status': status,
            'message': message,
            'estimated_size_mb': round(estimated_size_mb, 2),
            'usage_percent': round((estimated_size_mb / 1024) * 100, 1),
            'db_type': db_type,
            'db_icon': db_icon,
            'db_color': db_color,
            'stats': {
                'releves': nb_releves,
                'photos': nb_photos,
                'routines': nb_routines,
                'users': nb_users,
                'codes': nb_codes
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/database/cleanup', methods=['POST'])
@login_required
def api_database_cleanup():
    """Lance un nettoyage automatique de la base de données"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    try:
        success = cleanup_and_send_reports()
        if success:
            return jsonify({'message': 'Nettoyage effectué avec succès'})
        else:
            return jsonify({'error': 'Erreur lors du nettoyage'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/email/config', methods=['GET', 'PUT'])
@login_required
def api_email_config():
    """Gère la configuration email"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    if request.method == 'GET':
        config = get_email_config()
        return jsonify({
            'email_address': config.email_address,
            'smtp_server': config.smtp_server,
            'smtp_port': config.smtp_port,
            'smtp_username': config.smtp_username,
            'smtp_password': '***' if config.smtp_password else ''
        })
    
    elif request.method == 'PUT':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Données JSON manquantes'}), 400
        config = get_email_config()
        
        if 'email_address' in data:
            config.email_address = data['email_address']
        if 'smtp_server' in data:
            config.smtp_server = data['smtp_server']
        if 'smtp_port' in data:
            config.smtp_port = data['smtp_port']
        if 'smtp_username' in data:
            config.smtp_username = data['smtp_username']
        if 'smtp_password' in data and data['smtp_password'] != '***':
            config.smtp_password = data['smtp_password']
        
        db.session.commit()
        return jsonify({'message': 'Configuration email mise à jour'})

@app.route('/api/email/test', methods=['POST'])
@login_required
def api_test_email():
    """Teste la configuration email"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Accès non autorisé'}), 403
    
    try:
        config = get_email_config()
        if not config.email_address:
            return jsonify({'error': 'Aucune adresse email configurée'}), 400
        
        # Créer un email de test
        subject = "Test Configuration Email - STE Relevé"
        body = f"""
        <h2>Test de configuration email</h2>
        <p>Cet email confirme que la configuration email fonctionne correctement.</p>
        <p><strong>Date:</strong> {datetime.now().strftime('%d/%m/%Y à %H:%M')}</p>
        <p><strong>Serveur SMTP:</strong> {config.smtp_server}:{config.smtp_port}</p>
        """
        
        success = send_email_with_attachments(subject, body, [], config.email_address)
        
        if success:
            return jsonify({'message': 'Email de test envoyé avec succès'})
        else:
            return jsonify({'error': 'Erreur lors de l\'envoi de l\'email de test'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Modèle pour la configuration email
class CodeMagasin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(4), nullable=False)
    date = db.Column(db.Date, nullable=False, unique=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def __repr__(self):
        return f'<CodeMagasin {self.code} - {self.date}>'

class EmailConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email_address = db.Column(db.String(200), nullable=False, default='admin@ste-releve.com')
    smtp_server = db.Column(db.String(100), default='smtp.gmail.com')
    smtp_port = db.Column(db.Integer, default=587)
    smtp_username = db.Column(db.String(200))
    smtp_password = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Modèles pour la gestion du personnel
class Personnel(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, unique=True)
    nom = db.Column(db.String(100), nullable=False)
    prenom = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(200))
    telephone = db.Column(db.String(20))
    date_embauche = db.Column(db.Date)
    poste = db.Column(db.String(100))
    societe = db.Column(db.String(100))  # Société d'appartenance
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'), nullable=True)  # 1=SMP, 2=LPZ
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    user = relationship('User', foreign_keys=[user_id], back_populates='personnel')
    working_days = relationship('WorkingDays', back_populates='personnel', cascade='all, delete-orphan')
    leave_requests = relationship('LeaveRequest', back_populates='personnel', cascade='all, delete-orphan')
    documents = relationship('PersonnelDocument', back_populates='personnel', cascade='all, delete-orphan')
    absences = relationship('Absence', back_populates='personnel', cascade='all, delete-orphan')
    formations = relationship('Formation', back_populates='personnel', cascade='all, delete-orphan')

class WorkingDays(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=False)
    jour_semaine = db.Column(db.Integer, nullable=False)  # 0=lundi, 6=dimanche
    type_journee = db.Column(db.String(20), nullable=False)  # matin, apres_midi, journee, nuit
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    personnel = relationship('Personnel', back_populates='working_days')

class LeaveRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=False)
    date_debut = db.Column(db.Date, nullable=False)
    date_fin = db.Column(db.Date, nullable=False)
    type_conge = db.Column(db.String(50), nullable=False)  # conge_paye, conge_sans_solde, etc.
    statut = db.Column(db.String(20), default='en_attente')  # en_attente, accepte, refuse
    commentaire = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    personnel = relationship('Personnel', back_populates='leave_requests')
    documents = relationship('LeaveRequestDocument', back_populates='leave_request', cascade='all, delete-orphan')

# Modèle pour stocker la signature du manager
class ManagerSignature(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, unique=True)
    signature_path = db.Column(db.String(500), nullable=False)  # Chemin vers l'image de signature
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    user = relationship('User')

# Modèle pour stocker les PDFs de congé générés
class LeaveRequestDocument(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    leave_request_id = db.Column(db.Integer, db.ForeignKey('leave_request.id'), nullable=False)
    nom_fichier = db.Column(db.String(255), nullable=False)
    chemin_fichier = db.Column(db.String(500), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    leave_request = relationship('LeaveRequest', back_populates='documents')

class PersonnelDocument(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=False)
    nom_fichier = db.Column(db.String(255), nullable=False)
    chemin_fichier = db.Column(db.String(500), nullable=False)
    type_document = db.Column(db.String(100))  # contrat, cv, certificat, etc.
    description = db.Column(db.Text)
    uploaded_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    personnel = relationship('Personnel', back_populates='documents')
    uploader = relationship('User', foreign_keys=[uploaded_by])

class Absence(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=False)
    date_debut = db.Column(db.Date, nullable=False)
    date_fin = db.Column(db.Date, nullable=False)
    type_absence = db.Column(db.String(50), nullable=False)  # justifiee, injustifiee, formation
    statut = db.Column(db.String(20), default='en_attente')  # en_attente, validee, refusee
    commentaire = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    personnel = relationship('Personnel', back_populates='absences')
    creator = relationship('User', foreign_keys=[created_by])

class Formation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=False)
    nom_formation = db.Column(db.String(200), nullable=False)
    type_formation = db.Column(db.String(20), default='demande')  # demande, formation_obtenue
    date_debut = db.Column(db.Date, nullable=False)
    date_fin = db.Column(db.Date, nullable=False)
    date_fin_validite = db.Column(db.Date, nullable=True)  # Pour les formations obtenues
    statut = db.Column(db.String(20), default='prevue')  # prevue, realisee
    description = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    personnel = relationship('Personnel', back_populates='formations')
    creator = relationship('User', foreign_keys=[created_by])
    documents = relationship('FormationDocument', back_populates='formation', cascade='all, delete-orphan')

class FormationDocument(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    formation_id = db.Column(db.Integer, db.ForeignKey('formation.id'), nullable=False)
    nom_fichier = db.Column(db.String(255), nullable=False)
    chemin_fichier = db.Column(db.String(500), nullable=False)
    type_document = db.Column(db.String(100))  # certificat, attestation, etc.
    description = db.Column(db.Text)
    uploaded_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    formation = relationship('Formation', back_populates='documents')
    uploader = relationship('User', foreign_keys=[uploaded_by])

# Fonctions d'envoi d'email
def get_email_config():
    """Récupère la configuration email"""
    config = EmailConfig.query.first()
    if not config:
        config = EmailConfig()
        db.session.add(config)
        db.session.commit()
    return config

def send_email_with_attachments(subject, body, attachments, recipient_email):
    """Envoie un email avec pièces jointes (ZIP/PDF/images)"""
    try:
        config = get_email_config()
        
        # Créer le message
        msg = MIMEMultipart()
        msg['From'] = config.smtp_username or 'noreply@ste-releve.com'
        msg['To'] = recipient_email
        msg['Subject'] = subject
        
        # Corps du message
        msg.attach(MIMEText(body, 'html'))
        
        # Ajouter les pièces jointes avec le bon type MIME
        for attachment in attachments:
            with open(attachment['path'], 'rb') as f:
                filename = attachment['filename']
                if filename.lower().endswith('.zip'):
                    part = MIMEBase('application', 'zip')
                elif filename.lower().endswith('.pdf'):
                    part = MIMEBase('application', 'pdf')
                elif filename.lower().endswith('.jpg') or filename.lower().endswith('.jpeg'):
                    part = MIMEBase('image', 'jpeg')
                elif filename.lower().endswith('.png'):
                    part = MIMEBase('image', 'png')
                else:
                    part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                msg.attach(part)
        
        # Envoyer l'email
        if config.smtp_username and config.smtp_password:
            server = smtplib.SMTP(config.smtp_server, config.smtp_port)
            server.starttls()
            server.login(config.smtp_username, config.smtp_password)
            server.send_message(msg)
            server.quit()
            print(f"Email envoyé avec succès à {recipient_email}")
            return True
        else:
            print("Configuration SMTP manquante")
            return False
            
    except Exception as e:
        print(f"Erreur lors de l'envoi d'email: {e}")
        return False

def create_releve_20_zip(session_id):
    """Crée un fichier ZIP avec toutes les photos d'un relevé du 20 (toujours, même s'il n'y en a qu'une)"""
    try:
        photos = PhotoReleve.query.filter_by(session_id=session_id).all()
        if not photos:
            return None
        
        # Créer le ZIP
        zip_filename = f"releve_20_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        zip_path = os.path.join(app.config['UPLOAD_FOLDER'], zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            photos_ajoutees = 0
            for photo in photos:
                # Essayer d'abord le fichier local, puis le contenu de la base
                photo_path = os.path.join(app.config['UPLOAD_FOLDER'], photo.fichier_photo)
                if os.path.exists(photo_path):
                    # Utiliser le fichier local s'il existe
                    arcname = photo.fichier_photo
                    zipf.write(photo_path, arcname)
                    photos_ajoutees += 1
                    print(f"✅ Photo ajoutée depuis fichier local: {photo.fichier_photo}")
                elif photo.contenu_photo:
                    # Utiliser le contenu de la base de données
                    arcname = photo.fichier_photo
                    zipf.writestr(arcname, photo.contenu_photo)
                    photos_ajoutees += 1
                    print(f"✅ Photo ajoutée depuis base de données: {photo.fichier_photo}")
                else:
                    print(f"⚠️ Photo non trouvée (ni fichier ni contenu): {photo.fichier_photo}")
            
            if photos_ajoutees == 0:
                print(f"❌ Aucune photo trouvée pour la session {session_id}")
                return None
        
        # Vérification du contenu du ZIP (debug)
        with zipfile.ZipFile(zip_path, 'r') as zipf:
            print(f"Contenu du ZIP {zip_filename} : {zipf.namelist()}")
        
        return {
            'path': zip_path,
            'filename': zip_filename,
            'photos_count': len(photos)
        }
    except Exception as e:
        print(f"Erreur création ZIP relevé 20: {e}")
        return None



def cleanup_and_send_reports():
    try:
        config = get_email_config()
        if not config.email_address:
            print("Aucune adresse email configurée")
            return False
        print("🔄 Début du nettoyage automatique avec envoi des relevés du 20...")
        
        # 1. Envoyer les relevés du 20 avec dossier daté
        sessions_photos = db.session.query(PhotoReleve.session_id).distinct().all()
        for (session_id,) in sessions_photos:
            zip_file = create_releve_20_zip(session_id)
            if zip_file:
                first_photo = PhotoReleve.query.filter_by(session_id=session_id).first()
                if first_photo:
                    site = Site.query.get(first_photo.site_id)
                    user = User.query.get(first_photo.utilisateur_id)
                    # Utiliser la date de prise des photos pour le nom du dossier
                    date_photos = first_photo.date.strftime('%Y-%m-%d')
                    subject = f"Relevé du 20 - {site.nom if site else 'Site'} - {date_photos}"
                    body = f"""
                    <h2>Relevé du 20 - {site.nom if site else 'Site'}</h2>
                    <p><strong>Date des photos:</strong> {date_photos}</p>
                    <p><strong>Utilisateur:</strong> {user.username if user else 'Inconnu'}</p>
                    <p><strong>Nombre de photos:</strong> {zip_file['photos_count']}</p>
                    <p>Ce fichier ZIP contient toutes les photos du relevé directement à la racine.</p>
                    """
                    print(f"[CLEANUP] Envoi mail relevé 20: {subject}")
                    success = send_email_with_attachments(subject, body, [zip_file], config.email_address)
                    if success:
                        print(f"[CLEANUP] Mail relevé 20 envoyé, suppression des photos et relevés...")
                        # Supprimer les fichiers photos (si ils existent encore) - optionnel maintenant
                        photos_to_delete = PhotoReleve.query.filter_by(session_id=session_id).all()
                        for photo in photos_to_delete:
                            try:
                                file_path = os.path.join(app.config['UPLOAD_FOLDER'], photo.fichier_photo)
                                if os.path.exists(file_path):
                                    os.remove(file_path)
                                    print(f"✅ Fichier photo supprimé: {photo.fichier_photo}")
                                else:
                                    print(f"ℹ️ Fichier photo non trouvé (normal sur Render): {photo.fichier_photo}")
                            except Exception as e:
                                print(f"⚠️ Erreur suppression fichier {photo.fichier_photo}: {e}")
                        # Supprimer les enregistrements photos ET relevés du 20 (contenu binaire inclus)
                        PhotoReleve.query.filter_by(session_id=session_id).delete()
                        print(f"✅ Relevé 20 {session_id} envoyé et supprimé (contenu base inclus)")
                    try:
                        os.remove(zip_file['path'])
                    except:
                        pass
        
        db.session.commit()
        print("✅ Nettoyage automatique terminé avec succès")
        return True
    except Exception as e:
        print(f"❌ Erreur lors du nettoyage automatique: {e}")
        db.session.rollback()
        return False

@app.route('/api/accueil/synthese')
@login_required
def api_accueil_synthese():
    today = datetime.now().date()
    # Relevé SMP
    smp_fait = db.session.query(Releve).filter(Releve.site == 'SMP', Releve.date == today).count() > 0
    # Relevé LPZ
    lpz_fait = db.session.query(Releve).filter(Releve.site == 'LPZ', Releve.date == today).count() > 0
    # Routines (tous les formulaires)
    routines = []
    formulaires = FormulaireRoutine.query.order_by(FormulaireRoutine.nom).all()
    for f in formulaires:
        fait = db.session.query(ReponseRoutine).filter(ReponseRoutine.formulaire_id == f.id, ReponseRoutine.date_creation == today).count() > 0
        routines.append({'nom': f.nom, 'fait': fait})
    return jsonify({'smp': smp_fait, 'lpz': lpz_fait, 'routines': routines})

@app.route('/api/accueil/exhaure')
@login_required
def api_accueil_exhaure():
    today = datetime.now().date()
    # On suppose que le type de relevé exhaure est 'Exhaure' et qu'il y a un champ valeur, site
    sites = ['SMP', 'LPZ']
    result = []
    for site in sites:
        total = db.session.query(func.sum(Releve.valeur)).filter(Releve.site == site, Releve.type == 'Exhaure', Releve.date == today).scalar() or 0
        piscines = round(total / 2500, 2)
        result.append({'nom': site, 'm3': int(total), 'piscines': piscines})
    return jsonify(result)

# Stockage simple des dates de reset régularité (en mémoire, à remplacer par table si besoin)
RESET_REGULARITE = {}

@app.route('/api/accueil/synthese_v2')
@login_required
def api_accueil_synthese_v2():
    today = datetime.now().date()
    # Relevés fixes
    releves = []
    for nom in ['Relevé SMP', 'Relevé LPZ']:
        site = 'SMP' if 'SMP' in nom else 'LPZ'
        fait = db.session.query(Releve).filter(Releve.site == site, Releve.date == today).count() > 0
        # Calcul régularité sur 30 jours (hors reset)
        reset = RESET_REGULARITE.get(('releve', nom))
        date_debut = reset if reset else (today - timedelta(days=29))
        total = db.session.query(Releve.date).filter(Releve.site == site, Releve.date >= date_debut).distinct().count()
        jours = (today - date_debut).days + 1
        regularite = int(100 * total / jours) if jours > 0 else 0
        releves.append({'nom': nom, 'fait': fait, 'regularite': regularite})
    # Routines fixes
    routines_noms = [
        'STE PRINCIPALE SMP', 'STE CAB SMP', 'STEP SMP',
        'STE PRINCIPALE LPZ', 'STE CAB LPZ', 'STEP LPZ'
    ]
    routines = []
    for nom in routines_noms:
        formulaire = FormulaireRoutine.query.filter_by(nom=nom).first()
        fait = False
        regularite = 0
        if formulaire:
            fait = db.session.query(ReponseRoutine).filter(ReponseRoutine.formulaire_id == formulaire.id, ReponseRoutine.date_creation == today).count() > 0
            reset = RESET_REGULARITE.get(('routine', nom))
            date_debut = reset if reset else (today - timedelta(days=29))
            jours = (today - date_debut).days + 1
            total = db.session.query(ReponseRoutine.date_creation).filter(ReponseRoutine.formulaire_id == formulaire.id, ReponseRoutine.date_creation >= date_debut).distinct().count()
            regularite = int(100 * total / jours) if jours > 0 else 0
        routines.append({'nom': nom, 'fait': fait, 'regularite': regularite})
    return jsonify({'releves': releves, 'routines': routines})

@app.route('/api/accueil/exhaure_v2')
@login_required
def api_accueil_exhaure_v2():
    today = datetime.now().date()
    sites = ['SMP', 'LPZ']
    result = []
    for site in sites:
        total = 0
        for type_ in ['Exhaure', 'Bassin d\'orage', 'Retour dessableur']:
            total += db.session.query(func.sum(Releve.valeur)).filter(Releve.site == site, Releve.type == type_, Releve.date == today).scalar() or 0
        piscines = round(total / 2500, 2)
        result.append({'nom': site, 'm3': int(total), 'piscines': piscines})
    return jsonify(result)

@app.route('/api/accueil/reset_regularite', methods=['POST'])
@login_required
def api_accueil_reset_regularite():
    data = request.get_json()
    type_ = data.get('type')
    nom = data.get('nom')
    if not type_ or not nom:
        return jsonify({'error': 'Paramètres manquants'}), 400
    RESET_REGULARITE[(type_, nom)] = datetime.now().date()
    return jsonify({'success': True})

@app.route('/force_reset_db')
def force_reset_db():
    """Route temporaire pour forcer la réinitialisation de la base"""
    try:
        from werkzeug.security import generate_password_hash
        
        with app.app_context():
            # Supprimer toutes les tables
            db.drop_all()
            db.create_all()
            
            # Créer les sites
            smp = Site(nom='SMP', description='Station de traitement des eaux SMP')
            lpz = Site(nom='LPZ', description='Station de traitement des eaux LPZ')
            db.session.add(smp)
            db.session.add(lpz)
            db.session.commit()
            
            # Créer l'utilisateur admin
            admin = User()
            admin.username = 'admin'
            admin.password_hash = generate_password_hash('admin123')
            admin.role = 'admin'
            db.session.add(admin)
            db.session.commit()
            
            return """
            <h1>✅ Base de données réinitialisée avec succès!</h1>
            <p>Vous pouvez maintenant vous connecter avec :</p>
            <ul>
                <li><strong>Username:</strong> admin</li>
                <li><strong>Password:</strong> admin123</li>
            </ul>
            <p><a href="/login">Aller à la page de connexion</a></p>
            """
    except Exception as e:
        return f"<h1>❌ Erreur: {e}</h1>"



@app.route('/api/routines/reponses/<int:formulaire_id>/<date>')
@login_required
def api_reponses_formulaire_date(formulaire_id, date):
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400
    
    reponses = db.session.query(ReponseRoutine, QuestionRoutine).join(
        QuestionRoutine, ReponseRoutine.question_id == QuestionRoutine.id
    ).filter(
        ReponseRoutine.formulaire_id == formulaire_id,
        ReponseRoutine.date_creation == date_obj
    ).order_by(ReponseRoutine.heure_creation.asc()).all()
    
    result = []
    for reponse, question in reponses:
        result.append({
            'id': reponse.id,
            'formulaire_id': reponse.formulaire_id,
            'question_id': reponse.question_id,
            'id_question': question.id_question,
            'lieu': question.lieu,
            'question': question.question,
            'reponse': reponse.reponse,
            'commentaire': reponse.commentaire,
            'date_creation': reponse.date_creation.isoformat(),
            'heure_creation': reponse.heure_creation.isoformat() if reponse.heure_creation else None
        })
    
    return jsonify(result)

@app.route('/api/routines/reponses/<int:formulaire_id>/aujourdhui')
@login_required
def api_reponses_formulaire_aujourdhui(formulaire_id):
    """Récupère les réponses d'un formulaire pour aujourd'hui"""
    date_aujourdhui = datetime.now().date()
    
    reponses = db.session.query(ReponseRoutine, QuestionRoutine).join(
        QuestionRoutine, ReponseRoutine.question_id == QuestionRoutine.id
    ).filter(
        ReponseRoutine.formulaire_id == formulaire_id,
        ReponseRoutine.date_creation == date_aujourdhui
    ).order_by(ReponseRoutine.heure_creation.asc()).all()
    
    result = {}
    for reponse, question in reponses:
        result[question.id] = {
            'id': reponse.id,
            'formulaire_id': reponse.formulaire_id,
            'question_id': reponse.question_id,
            'id_question': question.id_question,
            'lieu': question.lieu,
            'question': question.question,
            'reponse': reponse.reponse,
            'commentaire': reponse.commentaire,
            'date_creation': reponse.date_creation.isoformat(),
            'heure_creation': reponse.heure_creation.isoformat() if reponse.heure_creation else None
        }
    
    return jsonify(result)

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000) 